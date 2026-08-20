import datetime as dt
import html
import json
import os
import re
import sys
import urllib.parse
import urllib.request
import http.cookiejar
import uuid
from pathlib import Path

from openpyxl import load_workbook, Workbook


BASE_URL = os.environ.get("XUI_BASE_URL", "http://xdplustv.online/99access99").rstrip("/")
XUI_USER = os.environ["XUI_USER"]
XUI_PASS = os.environ["XUI_PASS"]
INPUT_XLSX = Path(os.environ["XUI_CANDIDATES_XLSX"])
TODAY = dt.date.fromisoformat(os.environ.get("ANALYSIS_DATE", dt.date.today().isoformat()))
LIMIT = int(os.environ.get("XUI_APPLY_LIMIT", "0") or "0")
DRY_RUN = os.environ.get("XUI_DRY_RUN", "0") == "1"

SUPPORTED_MONTHS = {1, 3, 6, 12, 18, 24}
PACKAGE_MAP = {
    (3, False, 1): "60",
    (3, True, 1): "61",
    (6, False, 1): "62",
    (6, True, 1): "63",
    (3, False, 3): "64",
    (3, True, 3): "65",
    (3, False, 6): "66",
    (3, True, 6): "67",
    (3, False, 12): "68",
    (3, True, 12): "69",
    (6, False, 12): "70",
    (6, True, 12): "71",
    (10, False, 1): "72",
    (10, True, 1): "73",
    (1, False, 1): "74",
    (1, True, 1): "75",
    (6, True, 6): "76",
    (6, False, 6): "77",
    (6, True, 3): "78",
    (6, False, 3): "79",
}


def strip_html(value):
    text = str(value or "")
    text = re.sub(r"<br\s*/?>", " ", text, flags=re.I)
    text = re.sub(r"<[^>]+>", " ", text)
    return re.sub(r"\s+", " ", html.unescape(text)).strip()


def parse_date(value):
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    m = re.search(r"(\d{4}-\d{2}-\d{2})", strip_html(value))
    return dt.date.fromisoformat(m.group(1)) if m else None


def form_value(form, name, default=""):
    for pattern in [
        r"<input[^>]*name=[\"']" + re.escape(name) + r"[\"'][^>]*value=[\"']([^\"']*)[\"'][^>]*>",
        r"<input[^>]*value=[\"']([^\"']*)[\"'][^>]*name=[\"']" + re.escape(name) + r"[\"'][^>]*>",
    ]:
        m = re.search(pattern, form, flags=re.I)
        if m:
            return html.unescape(m.group(1))
    return default


def textarea_value(form, name):
    m = re.search(r"<textarea[^>]*name=[\"']" + re.escape(name) + r"[\"'][^>]*>(.*?)</textarea>", form, re.I | re.S)
    return html.unescape(m.group(1)).strip() if m else ""


def selected_value(form, name, default=""):
    m = re.search(r"<select[^>]*name=[\"']" + re.escape(name) + r"[\"'][^>]*>(.*?)</select>", form, re.I | re.S)
    if not m:
        return default
    selected = re.search(r"<option([^>]*)selected([^>]*)>", m.group(1), re.I)
    if selected:
        v = re.search(r"value=[\"']([^\"']*)", selected.group(0), re.I)
        if v:
            return html.unescape(v.group(1))
    return default


def selected_multi_values(form, name):
    m = re.search(r"<select[^>]*name=[\"']" + re.escape(name) + r"[\"'][^>]*>(.*?)</select>", form, re.I | re.S)
    if not m:
        return []
    values = []
    for attrs, _txt in re.findall(r"<option([^>]*)>(.*?)</option>", m.group(1), re.I | re.S):
        if "selected" in attrs.lower():
            v = re.search(r"value=[\"']([^\"']*)", attrs, re.I)
            if v:
                values.append(html.unescape(v.group(1)))
    return values


def is_checked(form, name):
    m = re.search(r"<input[^>]*name=[\"']" + re.escape(name) + r"[\"'][^>]*>", form, flags=re.I)
    return bool(m and "checked" in m.group(0).lower())


def current_bouquets(page):
    pos = page.find('$("#submit_button").val("Save")')
    frag = page[pos : pos + 20000] if pos >= 0 else page
    return re.findall(r"name='bouquets_selected\[\]' value='(\d+)' checked", frag)


def encode_multipart(fields):
    boundary = "----FenixBoundary" + uuid.uuid4().hex
    body = bytearray()
    for name, value in fields:
        body.extend(f"--{boundary}\r\n".encode())
        body.extend(f'Content-Disposition: form-data; name="{name}"\r\n\r\n'.encode())
        body.extend(str(value).encode("utf-8"))
        body.extend(b"\r\n")
    body.extend(f"--{boundary}--\r\n".encode())
    return bytes(body), f"multipart/form-data; boundary={boundary}"


def login():
    cookies = http.cookiejar.CookieJar()
    opener = urllib.request.build_opener(urllib.request.HTTPCookieProcessor(cookies))
    payload = urllib.parse.urlencode({"username": XUI_USER, "password": XUI_PASS, "referrer": "dashboard", "login": "Login"}).encode()
    with opener.open(urllib.request.Request(f"{BASE_URL}/login", data=payload, method="POST"), timeout=30) as r:
        if r.status >= 400:
            raise RuntimeError(f"Login XUI fallo HTTP {r.status}")
    return opener


def fetch_line(opener, line_id):
    with opener.open(f"{BASE_URL}/line?id={urllib.parse.quote(str(line_id))}", timeout=30) as r:
        page = r.read().decode("utf-8", "replace")
    m = re.search(r'<form[^>]*data-parsley-validate=""[^>]*>(.*?)</form>', page, re.I | re.S)
    if not m:
        raise RuntimeError(f"No encontre formulario de linea {line_id}")
    return page, m.group(1)


def find_user(opener, username):
    params = {
        "draw": "1", "start": "0", "length": "10", "id": "lines", "filter": "", "reseller": "3652",
        "order[0][column]": "0", "order[0][dir]": "desc", "search[value]": username, "search[regex]": "false",
    }
    for i in range(12):
        params[f"columns[{i}][data]"] = str(i)
        params[f"columns[{i}][name]"] = ""
        params[f"columns[{i}][searchable]"] = "true"
        params[f"columns[{i}][orderable]"] = "true"
        params[f"columns[{i}][search][value]"] = ""
        params[f"columns[{i}][search][regex]"] = "false"
    payload = json.loads(opener.open(f"{BASE_URL}/table?{urllib.parse.urlencode(params)}", timeout=60).read().decode("utf-8", "replace"))
    matches = []
    for row in payload.get("data") or []:
        rid = strip_html(row[0])
        user = strip_html(row[1])
        exp = parse_date(row[9])
        if user.lower() == username.lower():
            matches.append({"xui_id": rid, "usuario": user, "vencimiento": exp.isoformat() if exp else ""})
    return matches


def month_delta(target):
    if target.day != TODAY.day or target <= TODAY:
        return None
    return (target.year - TODAY.year) * 12 + (target.month - TODAY.month)


def package_sequence(months):
    if months in (1, 3, 6, 12):
        return [months]
    if months == 18:
        return [12, 6]
    if months == 24:
        return [12, 12]
    return []


def package_for(form, months):
    orig = form_value(form, "orig_package")
    max_conn = int(form_value(form, "max_connections", "3") or 3)
    has_xxx = "Con XXX" in orig
    package_id = PACKAGE_MAP.get((max_conn, has_xxx, months))
    if not package_id:
        raise RuntimeError(f"No existe paquete para {max_conn} conexiones, {'Con XXX' if has_xxx else 'Sin XXX'}, {months} meses")
    return package_id, orig, max_conn


def post_line(opener, fields):
    body, ctype = encode_multipart(fields)
    raw = opener.open(
        urllib.request.Request(f"{BASE_URL}/post.php?action=line", data=body, method="POST", headers={"Content-Type": ctype, "X-Requested-With": "XMLHttpRequest"}),
        timeout=60,
    ).read().decode("utf-8", "replace")
    try:
        return json.loads(raw)
    except Exception:
        return {"raw": raw[:1000]}


def delete_line(opener, line_id):
    if DRY_RUN:
        return {"dry_run": True, "delete": line_id}
    raw = opener.open(f"{BASE_URL}/api?action=line&sub=delete&user_id={urllib.parse.quote(str(line_id))}", timeout=30).read().decode("utf-8", "replace")
    try:
        return json.loads(raw)
    except Exception:
        return {"raw": raw[:1000]}


def create_line(opener, snapshot, package_months, username):
    form = snapshot["form"]
    page = snapshot["page"]
    package_id, orig_package, max_conn = package_for(form, package_months)
    bouquets = snapshot["bouquets"]
    fields = [
        ("bouquets_selected", ""),
        ("username", username),
        ("password", snapshot["password"]),
        ("member_id", snapshot["member_id"]),
        ("orig_package", orig_package),
        ("package", package_id),
        ("package_cost", ""),
        ("package_duration", ""),
        ("max_connections", str(max_conn)),
        ("exp_date", ""),
        ("contact", snapshot["contact"]),
        ("reseller_notes", snapshot["notes"]),
    ]
    for value in selected_multi_values(form, "allowed_ips[]"):
        fields.append(("allowed_ips[]", value))
    for value in selected_multi_values(form, "allowed_ua[]"):
        fields.append(("allowed_ua[]", value))
    if is_checked(form, "bypass_ua"):
        fields.append(("bypass_ua", "on"))
    if is_checked(form, "is_isplock"):
        fields.append(("is_isplock", "on"))
    fields.append(("isp_clear", form_value(form, "isp_clear")))
    for b in bouquets:
        fields.append(("bouquets_selected[]", b))
    fields.append(("submit_line", "Purchase"))
    if DRY_RUN:
        return {"dry_run": True, "package_id": package_id, "months": package_months}
    return post_line(opener, fields)


def load_candidates():
    wb = load_workbook(INPUT_XLSX, read_only=True, data_only=True)
    ws = wb["Acreditar"]
    rows = ws.iter_rows(values_only=True)
    headers = [str(x or "") for x in next(rows)]
    idx = {h: i for i, h in enumerate(headers)}
    out = []
    for row in rows:
        if row[idx["accion_sugerida"]] != "ACREDITAR":
            continue
        target = parse_date(row[idx["fenix_vencimiento"]])
        months = month_delta(target) if target else None
        if months not in SUPPORTED_MONTHS:
            continue
        out.append({
            "xui_id": str(row[idx["xui_id"]]).strip(),
            "xui_usuario": str(row[idx["xui_usuario"]]).strip(),
            "xui_vencimiento": str(row[idx["xui_vencimiento"]]).strip(),
            "fenix_vencimiento": target.isoformat(),
            "months": months,
        })
    return out[:LIMIT] if LIMIT else out


def save_log(results):
    out = Path("outputs") / f"xui_recreacion_fechas_fenix_{TODAY.isoformat()}.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Resultado"
    fields = ["xui_id_anterior", "xui_usuario", "xui_vencimiento_anterior", "fenix_vencimiento", "months", "xui_id_nuevo", "vencimiento_nuevo", "status", "detalle"]
    ws.append(fields)
    for r in results:
        ws.append([r.get(k, "") for k in fields])
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return out.resolve()


def main():
    opener = login()
    candidates = load_candidates()
    results = []
    for i, c in enumerate(candidates, 1):
        try:
            current_matches = find_user(opener, c["xui_usuario"])
            if current_matches:
                current = current_matches[0]
                if current.get("vencimiento") == c["fenix_vencimiento"]:
                    results.append({**c, "xui_id_anterior": c["xui_id"], "xui_id_nuevo": current.get("xui_id", ""), "vencimiento_nuevo": current.get("vencimiento", ""), "status": "OK_SKIP", "detalle": "Ya estaba alineado con fecha Fenix; no se toco en XUI."})
                    print(f"{i}/{len(candidates)} {c['xui_usuario']} ya alineado -> {c['fenix_vencimiento']} OK_SKIP")
                    continue
                c["xui_id"] = current.get("xui_id") or c["xui_id"]
            page, form = fetch_line(opener, c["xui_id"])
            bouquets = current_bouquets(page)
            if not bouquets:
                raise RuntimeError("No encontre bouquets actuales; se omite para no crear sin listas")
            snapshot = {
                "page": page,
                "form": form,
                "password": form_value(form, "password"),
                "member_id": selected_value(form, "member_id", "3652"),
                "contact": form_value(form, "contact"),
                "notes": textarea_value(form, "reseller_notes"),
                "bouquets": bouquets,
            }
            delete_response = delete_line(opener, c["xui_id"])
            if not DRY_RUN and not delete_response.get("result"):
                raise RuntimeError(f"No se pudo eliminar linea anterior: {delete_response}")
            responses = [delete_response]
            for months in package_sequence(c["months"]):
                responses.append(create_line(opener, snapshot, months, c["xui_usuario"]))
                found_after_step = find_user(opener, c["xui_usuario"])
                if not found_after_step:
                    raise RuntimeError(f"No se encontro usuario recreado despues de paquete {months}: {responses[-1]}")
                snapshot["page"], snapshot["form"] = fetch_line(opener, found_after_step[0]["xui_id"])
                snapshot["bouquets"] = current_bouquets(snapshot["page"]) or snapshot["bouquets"]
            found = find_user(opener, c["xui_usuario"])
            new = found[0] if found else {}
            ok = DRY_RUN or new.get("vencimiento") == c["fenix_vencimiento"]
            results.append({**c, "xui_id_anterior": c["xui_id"], "xui_id_nuevo": new.get("xui_id", ""), "vencimiento_nuevo": new.get("vencimiento", ""), "status": "OK" if ok else "REVISION", "detalle": json.dumps(responses, ensure_ascii=False)[:800]})
            print(f"{i}/{len(candidates)} {c['xui_usuario']} {c['xui_vencimiento']} -> {c['fenix_vencimiento']} {'OK' if ok else 'REVISION'}")
        except Exception as exc:
            results.append({**c, "xui_id_anterior": c.get("xui_id", ""), "status": "ERROR", "detalle": str(exc)})
            print(f"{i}/{len(candidates)} {c.get('xui_usuario')} ERROR {exc}")
            if LIMIT == 1:
                break
    log = save_log(results)
    print(f"LOG={log}")
    ok_statuses = {"OK", "OK_SKIP"}
    bad = [r for r in results if r.get("status") not in ok_statuses]
    print(f"TOTAL={len(results)} OK={len(results)-len(bad)} ERRORS={len(bad)}")
    if bad:
        sys.exit(2)


if __name__ == "__main__":
    main()
