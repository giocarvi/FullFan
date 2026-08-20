import os
from pathlib import Path

from openpyxl import load_workbook, Workbook


INPUT = Path(os.environ.get("INPUT_XLSX", "outputs/daily/analisis_xui_fenix_alineacion_2026-08-19.xlsx"))
SHEET = os.environ.get("INPUT_SHEET", "Vencen dia 19")
OUTPUT = Path(os.environ.get("OUTPUT_XLSX", "outputs/daily/candidatos_aplicar_dia_19.xlsx"))
SKIP_ONLINE = os.environ.get("SKIP_ONLINE", "1") == "1"
ONLY_ACTIVE_FENIX = os.environ.get("ONLY_ACTIVE_FENIX", "1") == "1"
ONLY_FENIX_SOURCE = os.environ.get("ONLY_FENIX_SOURCE", "1") == "1"


def is_online(value):
    return "Online:" in str(value or "")


def main():
    wb = load_workbook(INPUT, read_only=True, data_only=True)
    ws = wb[SHEET]
    rows = ws.iter_rows(values_only=True)
    headers = [str(x or "") for x in next(rows)]
    idx = {h: i for i, h in enumerate(headers)}

    out_wb = Workbook()
    out_ws = out_wb.active
    out_ws.title = "Acreditar"
    out_headers = ["xui_id", "xui_usuario", "xui_vencimiento", "fenix_vencimiento", "accion_sugerida"]
    out_ws.append(out_headers)

    total = kept = skipped_online = skipped_same = skipped_inactive = skipped_source = 0
    seen_ids = set()
    for row in rows:
        if not row or not row[idx["xui_id"]]:
            continue
        total += 1
        if ONLY_FENIX_SOURCE and "fuente_dia" in idx and "f" not in str(row[idx["fuente_dia"]] or "").strip().lower():
            skipped_source += 1
            continue
        xui_id = str(row[idx["xui_id"]] or "").strip()
        if xui_id in seen_ids:
            continue
        seen_ids.add(xui_id)
        diff = row[idx["dias_fenix_menos_xui"]]
        if diff in ("", None, 0):
            skipped_same += 1
            continue
        if SKIP_ONLINE and is_online(row[idx["last_connection"]]):
            skipped_online += 1
            continue
        if ONLY_ACTIVE_FENIX and str(row[idx["estado_fenix"]] or "").strip().lower() != "activo":
            skipped_inactive += 1
            continue
        out_ws.append([
            xui_id,
            row[idx["usuario"]],
            row[idx["vencimiento_xui"]],
            row[idx["vencimiento_fenix"]],
            "ACREDITAR",
        ])
        kept += 1

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    out_wb.save(OUTPUT)
    print(f"INPUT={INPUT}")
    print(f"SHEET={SHEET}")
    print(f"OUTPUT={OUTPUT.resolve()}")
    print(f"TOTAL={total} KEPT={kept} SKIPPED_ONLINE={skipped_online} SKIPPED_SAME={skipped_same} SKIPPED_INACTIVE={skipped_inactive} SKIPPED_SOURCE={skipped_source}")


if __name__ == "__main__":
    main()
