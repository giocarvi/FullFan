from flask import Flask, render_template, request, jsonify, session, redirect, url_for, Response
import os
import hmac
import json
import time
import re
import unicodedata
from urllib import request as urlrequest
from urllib.error import HTTPError, URLError
from urllib.parse import urlencode
from datetime import date, datetime, timezone, timedelta
from functools import wraps
from werkzeug.security import generate_password_hash, check_password_hash

# Zona horaria Guatemala (UTC-6)
GT_TZ = timezone(timedelta(hours=-6))

def today_gt():
    """Fecha actual en hora Guatemala."""
    return datetime.now(GT_TZ).date()

def now_gt():
    """Fecha y hora actual en Guatemala."""
    return datetime.now(GT_TZ)

def normalize_phone(value):
    digits = re.sub(r'\D+', '', str(value or ''))
    if digits.startswith('00'):
        digits = digits[2:]
    if len(digits) == 8:
        return '502' + digits
    return digits

def phone_country_label(value):
    digits = normalize_phone(value)
    country_codes = [
        ('502', 'Guatemala'),
        ('503', 'El Salvador'),
        ('504', 'Honduras'),
        ('505', 'Nicaragua'),
        ('506', 'Costa Rica'),
        ('507', 'Panamá'),
        ('52', 'México'),
        ('1', 'Estados Unidos / Canadá'),
    ]
    for prefix, label in country_codes:
        if digits.startswith(prefix):
            return label
    return 'Internacional' if len(digits) > 8 else ''

def normalize_text_key(value):
    """Normaliza texto para comparar nombres sin tildes, símbolos o espacios raros."""
    text = str(value or '').strip().lower()
    text = unicodedata.normalize('NFKD', text)
    text = ''.join(ch for ch in text if not unicodedata.combining(ch))
    text = re.sub(r'[^a-z0-9\s]+', ' ', text)
    return re.sub(r'\s+', ' ', text).strip()

def normalize_username_base(value):
    """Agrupa usernames parecidos removiendo símbolos y sufijos numéricos comunes."""
    text = normalize_text_key(value).replace(' ', '')
    text = re.sub(r'\d+$', '', text)
    return text.strip()

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY') or os.environ.get('FLASK_SECRET_KEY')
if not app.secret_key:
    raise RuntimeError('SECRET_KEY environment variable is required')

PASSWORD_HASH_PREFIXES = ('scrypt:', 'pbkdf2:', 'argon2:')
CREDIT_COST_USD = float(os.environ.get('CREDIT_COST_USD', '1.25'))
USD_GTQ_RATE = float(os.environ.get('USD_GTQ_RATE', '7.80'))
MAXPLAYER_API_BASE = os.environ.get('MAXPLAYER_API_BASE', 'https://api.maxplayer.tv/v3/api/public').rstrip('/')
MAXPLAYER_API_TOKEN = os.environ.get('MAXPLAYER_API_TOKEN', '')
MAXPLAYER_DOMAIN_ID = os.environ.get('MAXPLAYER_DOMAIN_ID', '')

def hash_password(password):
    return generate_password_hash(password)

def is_password_hash(value):
    return isinstance(value, str) and value.startswith(PASSWORD_HASH_PREFIXES)

def verify_password(stored_password, candidate_password):
    """Acepta hashes nuevos y contraseñas legacy en texto plano durante la transición."""
    if not stored_password:
        return False
    if is_password_hash(stored_password):
        return check_password_hash(stored_password, candidate_password)
    return hmac.compare_digest(str(stored_password), str(candidate_password))

def maybe_upgrade_password_hash(cursor, username, stored_password, candidate_password):
    """Si el usuario aún tiene password legacy, lo migra a hash después de login/cambio válido."""
    if not is_password_hash(stored_password) and verify_password(stored_password, candidate_password):
        cursor.execute(qmark("UPDATE usuarios SET password=? WHERE username=?"),
                       (hash_password(candidate_password), username))

# ── BASE DE DATOS ─────────────────────────────────────────────────────────────
# Si existe DATABASE_URL (Railway PostgreSQL), lo usa.
# Si no, usa SQLite local.
def estimate_credits_from_amount(amount):
    """Estima créditos históricos desde los planes conocidos, sin alterar datos antiguos."""
    try:
        amount = float(amount or 0)
    except (TypeError, ValueError):
        return 0
    if amount <= 0:
        return 0
    known_plans = [
        (90, 1),
        (225, 3),
        (400, 6),
        (700, 12),
        (1000, 18),
        (1300, 24),
    ]
    nearest_amount, nearest_credits = min(known_plans, key=lambda p: abs(amount - p[0]))
    if abs(amount - nearest_amount) > 120:
        return max(1, round(amount / 90))
    return nearest_credits


class MaxPlayerError(Exception):
    pass

def maxplayer_configured():
    return bool(MAXPLAYER_API_TOKEN and MAXPLAYER_DOMAIN_ID)

def maxplayer_request(method, path, payload=None):
    if not maxplayer_configured():
        raise MaxPlayerError('Max Player no está configurado. Agrega MAXPLAYER_API_TOKEN y MAXPLAYER_DOMAIN_ID en Railway.')
    url = f"{MAXPLAYER_API_BASE}{path}"
    body = None
    headers = {
        'Api-Token': MAXPLAYER_API_TOKEN,
        'Accept': 'application/json, text/plain, */*',
        'Accept-Language': 'es-GT,es;q=0.9,en;q=0.8',
        'Cache-Control': 'no-cache',
        'Pragma': 'no-cache',
        'Origin': 'https://my.maxplayer.tv',
        'Referer': 'https://my.maxplayer.tv/',
        'User-Agent': os.environ.get(
            'MAXPLAYER_API_USER_AGENT',
            'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 '
            '(KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36'
        ),
    }
    if payload is not None:
        body = json.dumps(payload).encode('utf-8')
        headers['Content-Type'] = 'application/json'
    req = urlrequest.Request(url, data=body, method=method.upper(), headers=headers)
    try:
        with urlrequest.urlopen(req, timeout=20) as response:
            raw = response.read().decode('utf-8') or '{}'
            try:
                parsed = json.loads(raw)
            except json.JSONDecodeError:
                return {'raw': raw}
            if isinstance(parsed, dict) and parsed.get('success') == 0 and parsed.get('error'):
                raise MaxPlayerError(f"Max Player respondió error: {parsed.get('error')}")
            return parsed
    except HTTPError as exc:
        raw = exc.read().decode('utf-8', errors='replace')
        try:
            detail = json.loads(raw)
        except json.JSONDecodeError:
            detail = raw
        raise MaxPlayerError(f"Max Player respondió HTTP {exc.code}: {detail}") from exc
    except URLError as exc:
        raise MaxPlayerError(f"No se pudo conectar con Max Player: {exc.reason}") from exc
    except TimeoutError as exc:
        raise MaxPlayerError('Max Player tardó demasiado en responder.') from exc

def extract_maxplayer_user_id(response):
    """Intenta encontrar el id del usuario sin depender de una forma exacta de respuesta."""
    if not isinstance(response, dict):
        return None
    candidates = [
        response.get('id'),
        response.get('user_id'),
        response.get('customer_id'),
    ]
    for key in ('user', 'data', 'customer'):
        nested = response.get(key)
        if isinstance(nested, dict):
            candidates.extend([nested.get('id'), nested.get('user_id'), nested.get('customer_id')])
    for value in candidates:
        if value not in (None, ''):
            return str(value)
    return None

def walk_dicts(payload):
    if isinstance(payload, dict):
        yield payload
        for value in payload.values():
            yield from walk_dicts(value)
    elif isinstance(payload, list):
        for item in payload:
            yield from walk_dicts(item)

def iter_maxplayer_users(payload):
    """Recorre respuestas comunes de Get Users sin depender de una estructura exacta."""
    if isinstance(payload, list):
        for item in payload:
            if isinstance(item, dict):
                yield item
        return
    if not isinstance(payload, dict):
        return
    for key in ('users', 'customers', 'data', 'items', 'results'):
        value = payload.get(key)
        if isinstance(value, list):
            for item in value:
                if isinstance(item, dict):
                    yield item
        elif isinstance(value, dict):
            for item in iter_maxplayer_users(value):
                yield item
    for item in walk_dicts(payload):
        if item is not payload and extract_maxplayer_user_id(item):
            yield item

def get_maxplayer_users(params=None):
    path = '/users'
    if params:
        path = f"{path}?{urlencode(params)}"
    return maxplayer_request('GET', path)

def find_maxplayer_user_id(username):
    if not username:
        return None
    target = str(username).strip().lower()
    attempts = [
        None,
        {'search': username},
        {'username': username},
        {'q': username},
        {'iptv_user': username},
        {'iptv_username': username},
    ]
    seen = set()
    for params in attempts:
        key = tuple(sorted((params or {}).items()))
        if key in seen:
            continue
        seen.add(key)
        response = get_maxplayer_users(params)
        for user in iter_maxplayer_users(response):
            possible_names = [
                user.get('username'),
                user.get('name'),
                user.get('user_name'),
                user.get('iptv_user'),
                user.get('iptv_username'),
                user.get('email'),
                user.get('user_email'),
            ]
            if any(str(value).strip().lower() == target for value in possible_names if value not in (None, '')):
                return extract_maxplayer_user_id(user)
    return None

def create_maxplayer_user(username, iptv_user, iptv_pass, password='', fullname='', user_email=''):
    payload = {
        'domain_id': str(MAXPLAYER_DOMAIN_ID),
        'iptv_user': iptv_user,
        'iptv_pass': iptv_pass,
    }
    if username:
        payload['username'] = username
    if password or username:
        payload['password'] = password or username
    if fullname:
        payload['fullname'] = fullname
    if user_email:
        payload['user_email'] = user_email
    response = maxplayer_request('POST', '/users', payload)
    user_id = extract_maxplayer_user_id(response)
    if not user_id:
        for attempt in range(3):
            if attempt:
                time.sleep(1)
            user_id = find_maxplayer_user_id(username)
            if user_id:
                break
    if not user_id:
        raise MaxPlayerError('Max Player creó el usuario, pero no devolvió/permitió ubicar el user_id para guardar la sincronización. Intenta nuevamente en unos segundos.')
    return response, user_id

def delete_maxplayer_user(user_id):
    if not user_id:
        raise MaxPlayerError('No hay user_id de Max Player para eliminar.')
    return maxplayer_request('DELETE', f'/users/{user_id}')

def delete_maxplayer_user_devices(user_id):
    if not user_id:
        raise MaxPlayerError('No hay user_id de Max Player para liberar dispositivos.')
    return maxplayer_request('DELETE', f'/users/devices/{user_id}')

def purge_maxplayer_user(user_id):
    """Libera dispositivos y elimina el usuario antes de recrearlo."""
    if not user_id:
        return
    try:
        delete_maxplayer_user_devices(user_id)
    except MaxPlayerError as exc:
        if not is_maxplayer_not_found_error(exc):
            raise
    try:
        delete_maxplayer_user(user_id)
    except MaxPlayerError as exc:
        if not is_maxplayer_not_found_error(exc):
            raise
    time.sleep(2)

def is_maxplayer_not_found_error(error):
    text = str(error).lower()
    return 'user not found' in text or 'not found' in text

def is_maxplayer_exists_error(error):
    text = str(error).lower()
    return 'already exist' in text or 'already exists' in text

DATABASE_URL = os.environ.get('DATABASE_URL', '')

if DATABASE_URL:
    import psycopg2
    import psycopg2.extras
    PG = True
    # Railway a veces usa "postgres://" en vez de "postgresql://"
    if DATABASE_URL.startswith('postgres://'):
        DATABASE_URL = DATABASE_URL.replace('postgres://', 'postgresql://', 1)

    def get_db():
        conn = psycopg2.connect(DATABASE_URL)
        return conn

    def qmark(sql):
        """Convierte ? de SQLite a %s de PostgreSQL."""
        return sql.replace('?', '%s')
else:
    import sqlite3
    PG = False
    DB_PATH = os.path.join(os.path.dirname(__file__), 'database.db')

    def get_db():
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row
        return conn

    def qmark(sql):
        return sql


def fetchone(cursor):
    row = cursor.fetchone()
    if row is None:
        return None
    if PG:
        return dict(zip([d[0] for d in cursor.description], row))
    return dict(row)


def fetchall(cursor):
    rows = cursor.fetchall()
    if PG:
        cols = [d[0] for d in cursor.description]
        return [dict(zip(cols, r)) for r in rows]
    return [dict(r) for r in rows]


# ── INICIALIZAR DB ────────────────────────────────────────────────────────────
def init_db():
    conn = get_db()
    c = conn.cursor()
    migration_wa_message = (
        'Hola {nombre}, le saludamos de Fénix Digital TV 🔥\n\n'
        'Le informamos que Full Fan Digital TV ya no opera, debido a la situación que se dio en Guatemala a principios de junio. '
        'Ahora es parte del grupo internacional Fénix Digital TV: su entretenimiento, sin fronteras.\n\n'
        'Seguimos con la misma atención, experiencia de 7 años y hoy lanzamos un nuevo servidor con nuevas opciones 😉\n\n'
        'Le invitamos a ingresar a nuestro portal de clientes:\n'
        '{portal}\n\n'
        'Usuario: {usuario}\n'
        'Password: {password_portal}\n\n'
        'En el portal podrá ver apps recomendadas, su usuario y password de Max Player, fecha de vencimiento e historial de pagos.\n\n'
        'Para Smart One el cambio se hará automático.\n\n'
        'Su servicio vence el {fecha}. Gracias por seguir con nosotros 🖐🏻'
    )

    if PG:
        c.execute('''CREATE TABLE IF NOT EXISTS clientes (
            username TEXT PRIMARY KEY,
            nombre TEXT,
            contacto TEXT,
            vencimiento TEXT,
            referido TEXT DEFAULT 'NO',
            parent_username TEXT DEFAULT NULL,
            total_pagado REAL DEFAULT 0,
            notas TEXT DEFAULT '',
            created_at TEXT DEFAULT (NOW()::text)
        )''')
        c.execute('''CREATE TABLE IF NOT EXISTS pagos (
            id SERIAL PRIMARY KEY,
            username TEXT,
            mes TEXT,
            monto REAL,
            fecha_registro TEXT DEFAULT (NOW()::text)
        )''')
        c.execute('''CREATE TABLE IF NOT EXISTS usuarios (
            id SERIAL PRIMARY KEY,
            username TEXT UNIQUE,
            password TEXT,
            rol TEXT DEFAULT 'atencion'
        )''')
        default_users = [
            ('admin', os.environ.get('DEFAULT_ADMIN_PASSWORD'), 'admin'),
            ('atencion', os.environ.get('DEFAULT_ATENCION_PASSWORD'), 'atencion'),
            ('jackye', os.environ.get('DEFAULT_JACKYE_PASSWORD'), 'atencion'),
            ('ingrid', os.environ.get('DEFAULT_INGRID_PASSWORD'), 'atencion'),
            ('turcios', os.environ.get('DEFAULT_TURCIOS_PASSWORD'), 'atencion'),
        ]
        for username, password, rol in default_users:
            if not password:
                continue
            c.execute("INSERT INTO usuarios (username,password,rol) VALUES (%s,%s,%s) ON CONFLICT DO NOTHING",
                      (username, hash_password(password), rol))
    else:
        c.executescript('''
            CREATE TABLE IF NOT EXISTS clientes (
                username TEXT PRIMARY KEY, nombre TEXT, contacto TEXT,
                vencimiento TEXT, referido TEXT DEFAULT 'NO',
                parent_username TEXT DEFAULT NULL,
                total_pagado REAL DEFAULT 0, notas TEXT DEFAULT '',
                created_at TEXT DEFAULT (datetime('now'))
            );
            CREATE TABLE IF NOT EXISTS pagos (
                id INTEGER PRIMARY KEY AUTOINCREMENT, username TEXT,
                mes TEXT, monto REAL,
                fecha_registro TEXT DEFAULT (datetime('now'))
            );
            CREATE TABLE IF NOT EXISTS usuarios (
                id INTEGER PRIMARY KEY AUTOINCREMENT, username TEXT UNIQUE,
                password TEXT, rol TEXT DEFAULT 'atencion'
            );
        ''')
        default_users = [
            ('admin', os.environ.get('DEFAULT_ADMIN_PASSWORD'), 'admin'),
            ('atencion', os.environ.get('DEFAULT_ATENCION_PASSWORD'), 'atencion'),
            ('jackye', os.environ.get('DEFAULT_JACKYE_PASSWORD'), 'atencion'),
            ('ingrid', os.environ.get('DEFAULT_INGRID_PASSWORD'), 'atencion'),
            ('turcios', os.environ.get('DEFAULT_TURCIOS_PASSWORD'), 'atencion'),
        ]
        for username, password, rol in default_users:
            if not password:
                continue
            c.execute("INSERT OR IGNORE INTO usuarios (username,password,rol) VALUES (?,?,?)",
                      (username, hash_password(password), rol))

    # Migrar: agregar columna comprobante a pagos si no existe
    if PG:
        c.execute("ALTER TABLE pagos ADD COLUMN IF NOT EXISTS comprobante TEXT DEFAULT NULL")
        c.execute("ALTER TABLE pagos ADD COLUMN IF NOT EXISTS created_by TEXT DEFAULT NULL")
        c.execute("ALTER TABLE clientes ADD COLUMN IF NOT EXISTS contacto_secundario TEXT DEFAULT NULL")
        c.execute("ALTER TABLE clientes ADD COLUMN IF NOT EXISTS email TEXT DEFAULT NULL")
        c.execute("ALTER TABLE clientes ADD COLUMN IF NOT EXISTS datos_actualizados_at TEXT DEFAULT NULL")
        c.execute("ALTER TABLE clientes ADD COLUMN IF NOT EXISTS parent_username TEXT DEFAULT NULL")
    else:
        try:
            c.execute("ALTER TABLE pagos ADD COLUMN comprobante TEXT DEFAULT NULL")
        except Exception:
            pass
        try:
            c.execute("ALTER TABLE pagos ADD COLUMN created_by TEXT DEFAULT NULL")
        except Exception:
            pass
        for ddl in (
            "ALTER TABLE clientes ADD COLUMN contacto_secundario TEXT DEFAULT NULL",
            "ALTER TABLE clientes ADD COLUMN email TEXT DEFAULT NULL",
            "ALTER TABLE clientes ADD COLUMN datos_actualizados_at TEXT DEFAULT NULL",
            "ALTER TABLE clientes ADD COLUMN parent_username TEXT DEFAULT NULL",
        ):
            try:
                c.execute(ddl)
            except Exception:
                pass

    # Tabla de configuración
    if PG:
        c.execute('''CREATE TABLE IF NOT EXISTS configuracion (
            clave TEXT PRIMARY KEY,
            valor TEXT
        )''')
        defaults = [
            ('wa_prefijo', '502'),
            ('wa_saludo', 'Hola {nombre}, te saludamos de Fénix Digital TV 👋 Tu entretenimiento, sin fronteras.'),
            ('wa_recordatorio', 'Hola {nombre}, tu servicio de Fénix Digital TV vence el {fecha}. Puedes renovar antes de la fecha para evitar interrupciones. ¡Gracias! 🔥'),
            ('wa_confirmar_pago', 'Hola {nombre}, hemos recibido tu pago ✅. Tu servicio Fénix Digital TV ha sido renovado hasta el {fecha}. ¡Gracias por preferirnos! 🔥'),
            ('wa_vencido', 'Hola {nombre}, tu servicio de Fénix Digital TV ha vencido 📅. Para reactivarlo realiza tu pago y envíanos el comprobante. ¡Te esperamos! 🔥'),
            ('wa_migracion', migration_wa_message),
        ]
        for clave, valor in defaults:
            c.execute("INSERT INTO configuracion (clave, valor) VALUES (%s, %s) ON CONFLICT DO NOTHING", (clave, valor))
    else:
        c.execute('''CREATE TABLE IF NOT EXISTS configuracion (
            clave TEXT PRIMARY KEY,
            valor TEXT
        )''')
        defaults = [
            ('wa_prefijo', '502'),
            ('wa_saludo', 'Hola {nombre}, te saludamos de Fénix Digital TV 👋 Tu entretenimiento, sin fronteras.'),
            ('wa_recordatorio', 'Hola {nombre}, tu servicio de Fénix Digital TV vence el {fecha}. Puedes renovar antes de la fecha para evitar interrupciones. ¡Gracias! 🔥'),
            ('wa_confirmar_pago', 'Hola {nombre}, hemos recibido tu pago ✅. Tu servicio Fénix Digital TV ha sido renovado hasta el {fecha}. ¡Gracias por preferirnos! 🔥'),
            ('wa_vencido', 'Hola {nombre}, tu servicio de Fénix Digital TV ha vencido 📅. Para reactivarlo realiza tu pago y envíanos el comprobante. ¡Te esperamos! 🔥'),
            ('wa_migracion', migration_wa_message),
        ]
        for clave, valor in defaults:
            c.execute("INSERT OR IGNORE INTO configuracion (clave, valor) VALUES (?, ?)", (clave, valor))

    # Rebrand suave: solo reemplaza plantillas antiguas que aún mencionen Full Fan.
    brand_updates = [
        ('wa_saludo', 'Hola {nombre}, te saludamos de Fénix Digital TV 👋 Tu entretenimiento, sin fronteras.'),
        ('wa_recordatorio', 'Hola {nombre}, tu servicio de Fénix Digital TV vence el {fecha}. Puedes renovar antes de la fecha para evitar interrupciones. ¡Gracias! 🔥'),
        ('wa_confirmar_pago', 'Hola {nombre}, hemos recibido tu pago ✅. Tu servicio Fénix Digital TV ha sido renovado hasta el {fecha}. ¡Gracias por preferirnos! 🔥'),
        ('wa_vencido', 'Hola {nombre}, tu servicio de Fénix Digital TV ha vencido 📅. Para reactivarlo realiza tu pago y envíanos el comprobante. ¡Te esperamos! 🔥'),
        ('wa_migracion', migration_wa_message),
    ]
    for clave, valor in brand_updates:
        c.execute(qmark("UPDATE configuracion SET valor=? WHERE clave=? AND valor LIKE ?"),
                  (valor, clave, '%Full Fan%'))

    # ── FÉNIX OPERACIÓN: planes, pedidos y cola manual ────────────────────────
    if PG:
        c.execute('''CREATE TABLE IF NOT EXISTS plans (
            id SERIAL PRIMARY KEY,
            slug TEXT UNIQUE NOT NULL,
            name TEXT NOT NULL,
            months INTEGER NOT NULL,
            credits_required INTEGER NOT NULL,
            connections INTEGER NOT NULL DEFAULT 3,
            price_gtq REAL,
            price_usd REAL,
            is_active BOOLEAN NOT NULL DEFAULT TRUE,
            created_at TEXT DEFAULT (NOW()::text)
        )''')
        c.execute('''CREATE TABLE IF NOT EXISTS orders (
            id SERIAL PRIMARY KEY,
            username TEXT,
            plan_id INTEGER,
            type TEXT DEFAULT 'renewal',
            status TEXT DEFAULT 'pending_activation',
            amount REAL DEFAULT 0,
            currency TEXT DEFAULT 'GTQ',
            credits_required INTEGER DEFAULT 0,
            payment_method TEXT,
            payment_proof TEXT,
            payment_registered_at TEXT,
            payment_id INTEGER,
            notes TEXT,
            created_at TEXT DEFAULT (NOW()::text),
            completed_at TEXT
        )''')
        c.execute("ALTER TABLE orders ADD COLUMN IF NOT EXISTS payment_proof TEXT DEFAULT NULL")
        c.execute("ALTER TABLE orders ADD COLUMN IF NOT EXISTS payment_registered_at TEXT DEFAULT NULL")
        c.execute("ALTER TABLE orders ADD COLUMN IF NOT EXISTS payment_id INTEGER DEFAULT NULL")
        c.execute('''CREATE TABLE IF NOT EXISTS activation_tasks (
            id SERIAL PRIMARY KEY,
            order_id INTEGER,
            username TEXT,
            task_type TEXT DEFAULT 'renew_line',
            status TEXT DEFAULT 'pending',
            assigned_to TEXT,
            xui_username TEXT,
            xui_expires_at TEXT,
            credits_to_consume INTEGER DEFAULT 0,
            notes TEXT,
            blocked_reason TEXT,
            created_at TEXT DEFAULT (NOW()::text),
            completed_at TEXT
        )''')
        c.execute("ALTER TABLE activation_tasks ADD COLUMN IF NOT EXISTS xui_password TEXT DEFAULT NULL")
        c.execute('''CREATE TABLE IF NOT EXISTS client_portal_accounts (
            username TEXT PRIMARY KEY,
            password TEXT NOT NULL,
            is_enabled BOOLEAN NOT NULL DEFAULT TRUE,
            created_at TEXT DEFAULT (NOW()::text),
            updated_at TEXT
        )''')
        c.execute('''CREATE TABLE IF NOT EXISTS client_service_credentials (
            username TEXT PRIMARY KEY,
            app_name TEXT DEFAULT 'Max Player',
            service_username TEXT,
            service_password TEXT,
            expires_at TEXT,
            devices INTEGER DEFAULT 3,
            notes TEXT,
            maxplayer_user_id TEXT,
            maxplayer_synced_at TEXT,
            maxplayer_sync_status TEXT,
            updated_at TEXT DEFAULT (NOW()::text)
        )''')
        c.execute("ALTER TABLE client_service_credentials ADD COLUMN IF NOT EXISTS maxplayer_user_id TEXT DEFAULT NULL")
        c.execute("ALTER TABLE client_service_credentials ADD COLUMN IF NOT EXISTS maxplayer_synced_at TEXT DEFAULT NULL")
        c.execute("ALTER TABLE client_service_credentials ADD COLUMN IF NOT EXISTS maxplayer_sync_status TEXT DEFAULT NULL")
        c.execute('''CREATE TABLE IF NOT EXISTS device_apps (
            id SERIAL PRIMARY KEY,
            device_type TEXT NOT NULL,
            app_name TEXT NOT NULL,
            guide_slug TEXT,
            is_active BOOLEAN NOT NULL DEFAULT TRUE,
            created_at TEXT DEFAULT (NOW()::text)
        )''')
        plan_defaults = [
            ('fundador-1m', 'Fundador 1 mes', 1, 1, 3, 90, None),
            ('fundador-3m', 'Fundador 3 meses', 3, 3, 3, 225, None),
            ('fundador-6m', 'Fundador 6 meses', 6, 6, 3, 400, None),
            ('fundador-12m', 'Fundador 12 meses', 12, 12, 3, 700, None),
            ('familiar-1m', 'Familiar 1 mes', 1, 1, 3, None, 13.99),
            ('familiar-3m', 'Familiar 3 meses', 3, 3, 3, None, 38.99),
            ('familiar-6m', 'Familiar 6 meses', 6, 6, 3, None, 73.99),
            ('familiar-12m', 'Familiar 12 meses', 12, 12, 3, None, 137.99),
        ]
        for p in plan_defaults:
            c.execute("""INSERT INTO plans (slug,name,months,credits_required,connections,price_gtq,price_usd)
                         VALUES (%s,%s,%s,%s,%s,%s,%s) ON CONFLICT (slug) DO NOTHING""", p)
        device_defaults = [
            ('Android', 'Max Player', 'android'),
            ('iOS', 'Max Player', 'ios'),
            ('Firestick', 'Max Player', 'firestick'),
            ('Samsung Smart TV', 'Max Player', 'samsung'),
            ('LG Smart TV', 'Max Player', 'lg'),
            ('Windows', 'Max Player', 'windows'),
            ('Mac', 'Max Player', 'mac'),
            ('Hisense Smart TV', 'Smart One', 'hisense'),
            ('Roku', 'Premium Player', 'roku'),
        ]
        for d in device_defaults:
            c.execute("""INSERT INTO device_apps (device_type,app_name,guide_slug)
                         SELECT %s,%s,%s WHERE NOT EXISTS (
                           SELECT 1 FROM device_apps WHERE device_type=%s AND app_name=%s
                         )""", d + (d[0], d[1]))
    else:
        c.executescript('''
            CREATE TABLE IF NOT EXISTS plans (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                slug TEXT UNIQUE NOT NULL,
                name TEXT NOT NULL,
                months INTEGER NOT NULL,
                credits_required INTEGER NOT NULL,
                connections INTEGER NOT NULL DEFAULT 3,
                price_gtq REAL,
                price_usd REAL,
                is_active INTEGER NOT NULL DEFAULT 1,
                created_at TEXT DEFAULT (datetime('now'))
            );
            CREATE TABLE IF NOT EXISTS orders (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT,
                plan_id INTEGER,
                type TEXT DEFAULT 'renewal',
                status TEXT DEFAULT 'pending_activation',
                amount REAL DEFAULT 0,
                currency TEXT DEFAULT 'GTQ',
                credits_required INTEGER DEFAULT 0,
                payment_method TEXT,
                payment_proof TEXT,
                payment_registered_at TEXT,
                payment_id INTEGER,
                notes TEXT,
                created_at TEXT DEFAULT (datetime('now')),
                completed_at TEXT
            );
            CREATE TABLE IF NOT EXISTS activation_tasks (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                order_id INTEGER,
                username TEXT,
                task_type TEXT DEFAULT 'renew_line',
                status TEXT DEFAULT 'pending',
                assigned_to TEXT,
                xui_username TEXT,
                xui_expires_at TEXT,
                credits_to_consume INTEGER DEFAULT 0,
                notes TEXT,
                blocked_reason TEXT,
                created_at TEXT DEFAULT (datetime('now')),
                completed_at TEXT
            );
            CREATE TABLE IF NOT EXISTS client_portal_accounts (
                username TEXT PRIMARY KEY,
                password TEXT NOT NULL,
                is_enabled INTEGER NOT NULL DEFAULT 1,
                created_at TEXT DEFAULT (datetime('now')),
                updated_at TEXT
            );
            CREATE TABLE IF NOT EXISTS client_service_credentials (
                username TEXT PRIMARY KEY,
                app_name TEXT DEFAULT 'Max Player',
                service_username TEXT,
                service_password TEXT,
                expires_at TEXT,
                devices INTEGER DEFAULT 3,
                notes TEXT,
                maxplayer_user_id TEXT,
                maxplayer_synced_at TEXT,
                maxplayer_sync_status TEXT,
                updated_at TEXT DEFAULT (datetime('now'))
            );
            CREATE TABLE IF NOT EXISTS device_apps (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                device_type TEXT NOT NULL,
                app_name TEXT NOT NULL,
                guide_slug TEXT,
                is_active INTEGER NOT NULL DEFAULT 1,
                created_at TEXT DEFAULT (datetime('now'))
            );
        ''')
        plan_defaults = [
            ('fundador-1m', 'Fundador 1 mes', 1, 1, 3, 90, None),
            ('fundador-3m', 'Fundador 3 meses', 3, 3, 3, 225, None),
            ('fundador-6m', 'Fundador 6 meses', 6, 6, 3, 400, None),
            ('fundador-12m', 'Fundador 12 meses', 12, 12, 3, 700, None),
            ('familiar-1m', 'Familiar 1 mes', 1, 1, 3, None, 13.99),
            ('familiar-3m', 'Familiar 3 meses', 3, 3, 3, None, 38.99),
            ('familiar-6m', 'Familiar 6 meses', 6, 6, 3, None, 73.99),
            ('familiar-12m', 'Familiar 12 meses', 12, 12, 3, None, 137.99),
        ]
        for p in plan_defaults:
            c.execute("""INSERT OR IGNORE INTO plans
                         (slug,name,months,credits_required,connections,price_gtq,price_usd)
                         VALUES (?,?,?,?,?,?,?)""", p)
        device_defaults = [
            ('Android', 'Max Player', 'android'),
            ('iOS', 'Max Player', 'ios'),
            ('Firestick', 'Max Player', 'firestick'),
            ('Samsung Smart TV', 'Max Player', 'samsung'),
            ('LG Smart TV', 'Max Player', 'lg'),
            ('Windows', 'Max Player', 'windows'),
            ('Mac', 'Max Player', 'mac'),
            ('Hisense Smart TV', 'Smart One', 'hisense'),
            ('Roku', 'Premium Player', 'roku'),
        ]
        for d in device_defaults:
            c.execute("""INSERT INTO device_apps (device_type,app_name,guide_slug)
                         SELECT ?,?,? WHERE NOT EXISTS (
                           SELECT 1 FROM device_apps WHERE device_type=? AND app_name=?
                         )""", d + (d[0], d[1]))

        try:
            c.execute("ALTER TABLE orders ADD COLUMN payment_proof TEXT DEFAULT NULL")
        except Exception:
            pass
        try:
            c.execute("ALTER TABLE orders ADD COLUMN payment_registered_at TEXT DEFAULT NULL")
        except Exception:
            pass
        try:
            c.execute("ALTER TABLE orders ADD COLUMN payment_id INTEGER DEFAULT NULL")
        except Exception:
            pass
        try:
            c.execute("ALTER TABLE activation_tasks ADD COLUMN xui_password TEXT DEFAULT NULL")
        except Exception:
            pass
        for ddl in (
            "ALTER TABLE client_service_credentials ADD COLUMN maxplayer_user_id TEXT DEFAULT NULL",
            "ALTER TABLE client_service_credentials ADD COLUMN maxplayer_synced_at TEXT DEFAULT NULL",
            "ALTER TABLE client_service_credentials ADD COLUMN maxplayer_sync_status TEXT DEFAULT NULL",
        ):
            try:
                c.execute(ddl)
            except Exception:
                pass

    conn.commit()

    # Migrar Excel si DB está vacía
    c.execute("SELECT COUNT(*) as c FROM clientes")
    total = c.fetchone()[0] if PG else fetchone(c)['c']
    conn.close()
    if total == 0:
        _migrate_from_excel()


def _parse_excel_rows(excel_path):
    """Lee el Excel con openpyxl y devuelve (headers_row, data_rows) como listas."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        import subprocess, sys
        subprocess.check_call([sys.executable, '-m', 'pip', 'install', '--quiet', 'openpyxl'])
        from openpyxl import load_workbook
    from datetime import datetime
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()
    headers = list(all_rows[1])   # fila 2 = índice 1
    data = all_rows[2:]           # desde fila 3
    return headers, data


def _cell_val(v):
    return v if v is not None else None


def _migrate_from_excel(update_existing=False):
    excel_path = os.path.join(os.path.dirname(__file__), 'IPTV Nuevo (2).xlsx')
    if not os.path.exists(excel_path):
        return 0, 0
    try:
        return _import_excel_rows(excel_path, update_existing)
    except Exception as e:
        print(f"Error migrando Excel: {e}")
        return 0, 0


def _import_excel_rows(excel_path, update_existing=False):
    from datetime import datetime as dt
    headers, data = _parse_excel_rows(excel_path)
    conn = get_db()
    c = conn.cursor()
    clients_ok = 0
    pagos_ok = 0
    for row in data:
        if not row or len(row) < 5:
            continue
        username = str(row[0]).strip() if row[0] is not None else None
        if not username or username.lower() == 'nan' or username == '':
            continue
        nombre = str(row[1]).strip() if row[1] is not None else ''
        exp_val = row[2]
        contact = str(row[3]).strip() if row[3] is not None else ''
        referido_raw = str(row[4]).strip().upper() if row[4] is not None else 'NO'
        referido = 'SI' if referido_raw in ('SI', 'S') else 'NO'
        total = float(row[82]) if len(row) > 82 and row[82] is not None else 0
        exp_str = None
        if exp_val is not None:
            try:
                if isinstance(exp_val, (dt,)):
                    exp_str = exp_val.strftime('%Y-%m-%d')
                else:
                    from datetime import datetime
                    exp_str = datetime.strptime(str(exp_val)[:10], '%Y-%m-%d').strftime('%Y-%m-%d')
            except:
                pass
        if PG:
            if update_existing:
                c.execute("""INSERT INTO clientes (username,nombre,contacto,vencimiento,referido,total_pagado)
                             VALUES (%s,%s,%s,%s,%s,%s)
                             ON CONFLICT (username) DO UPDATE SET
                               nombre=EXCLUDED.nombre, contacto=EXCLUDED.contacto,
                               vencimiento=EXCLUDED.vencimiento, referido=EXCLUDED.referido,
                               total_pagado=EXCLUDED.total_pagado""",
                          (username, nombre, contact, exp_str, referido, total))
            else:
                c.execute("INSERT INTO clientes (username,nombre,contacto,vencimiento,referido,total_pagado) VALUES (%s,%s,%s,%s,%s,%s) ON CONFLICT DO NOTHING",
                          (username, nombre, contact, exp_str, referido, total))
        else:
            c.execute("INSERT OR IGNORE INTO clientes (username,nombre,contacto,vencimiento,referido,total_pagado) VALUES (?,?,?,?,?,?)",
                      (username, nombre, contact, exp_str, referido, total))
        clients_ok += 1
        for col in range(5, min(82, len(row))):
            val = row[col]
            if val is not None and isinstance(val, (int, float)) and val > 0:
                month_date = headers[col] if col < len(headers) else None
                if month_date is not None and hasattr(month_date, 'strftime'):
                    month_str = month_date.strftime('%Y-%m-%d')
                    if PG:
                        c.execute("INSERT INTO pagos (username,mes,monto) SELECT %s,%s,%s WHERE NOT EXISTS (SELECT 1 FROM pagos WHERE username=%s AND mes=%s)",
                                  (username, month_str, float(val), username, month_str))
                    else:
                        ex = c.execute("SELECT 1 FROM pagos WHERE username=? AND mes=?", (username, month_str)).fetchone()
                        if not ex:
                            c.execute("INSERT INTO pagos (username,mes,monto) VALUES (?,?,?)", (username, month_str, float(val)))
                    pagos_ok += 1
    conn.commit()
    conn.close()
    print(f"Migración desde Excel completada: {clients_ok} clientes, {pagos_ok} pagos.")
    return clients_ok, pagos_ok


def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated


def client_login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'client_username' not in session:
            return redirect(url_for('client_login'))
        return f(*args, **kwargs)
    return decorated


def service_status(expires_at):
    today = today_gt().isoformat()
    if not expires_at:
        return 'pendiente'
    return 'activo' if str(expires_at) >= today else 'vencido'

# ── AUTH ──────────────────────────────────────────────────────────────────────
@app.route('/inicio')
def public_home():
    return render_template('public_home.html')


@app.route('/')
def public_root():
    return redirect(url_for('public_home'))


DEVICE_GUIDES = {
    'android-ios': {
        'title': 'Celulares/Tablets Android / iOS',
        'app': 'Max Player',
        'icon': '📱',
        'logo': 'apps/max-player.webp',
        'hero_image': 'landing/device-android-ios.webp',
        'steps': [
            'Abre la tienda de aplicaciones de tu dispositivo.',
            'Busca e instala Max Player.',
            'Escríbenos por WhatsApp para solicitar demo o activar tu plan.',
            'Ingresa el usuario y contraseña que te enviaremos.',
            'Prueba la reproducción y confirma que todo funcione correctamente.'
        ],
        'tips': ['Usa una conexión WiFi estable.', 'Si cambias de celular, avísanos para revisar tu acceso.'],
        'links': [
            {'label': 'Max Player para Android', 'url': 'https://play.google.com/store/apps/details?id=tv.maxplayer.android'},
            {'label': 'Max Player para iOS', 'url': 'https://apps.apple.com/app/maxplayer-iptv-player/id1660982028'},
            {'label': 'Sitio oficial Max Player', 'url': 'https://maxplayer.tv/en'},
        ],
        'app_options': [
            {
                'title': 'Opción 1 · Max Player',
                'subtitle': 'Disponible para celulares y tablets Android / iOS.',
                'logo': 'apps/max-player.webp',
                'steps': [
                    'Abre Google Play o App Store según tu dispositivo.',
                    'Busca e instala Max Player.',
                    'Ingresa el usuario y contraseña que te enviaremos.',
                    'Prueba la reproducción y confirma que todo funcione correctamente.'
                ],
                'links': [
                    {'label': 'Max Player para Android', 'url': 'https://play.google.com/store/apps/details?id=tv.maxplayer.android'},
                    {'label': 'Max Player para iOS', 'url': 'https://apps.apple.com/app/maxplayer-iptv-player/id1660982028'},
                    {'label': 'Sitio oficial Max Player', 'url': 'https://maxplayer.tv/en'},
                ]
            },
            {
                'title': 'Opción 2 · XD Plus',
                'subtitle': 'Únicamente para celulares y tablets Android.',
                'logo': 'apps/xd-plus.webp',
                'download_code': 'Downloader · código 8813174',
                'steps': [
                    'Descarga e instala la app Downloader en tu Android.',
                    'Abre Downloader y escribe el código 8813174.',
                    'Instala XD Plus y abre la aplicación.',
                    'Escríbenos por WhatsApp para confirmar tus datos de acceso.'
                ],
                'links': []
            }
        ]
    },
    'firestick': {
        'title': 'Amazon Firestick / Android Stick / TV Box Android',
        'app': 'Max Player / XD+',
        'icon': '🔥',
        'logo': 'apps/max-player.webp',
        'hero_image': 'landing/device-firestick.webp',
        'steps': [
            'Conecta tu Firestick a internet.',
            'Busca Max Player desde tu tienda o método de instalación disponible.',
            'También puedes instalar XD+ usando Downloader con el código 8813174.',
            'Instala la app y ábrela.',
            'Solicita tus datos de acceso por WhatsApp.',
            'Ingresa usuario y contraseña y prueba un canal.'
        ],
        'tips': ['Mantén espacio libre en el Firestick.', 'Si usas Downloader, confirma que el código sea 8813174.', 'Reinicia el dispositivo si una app queda congelada.'],
        'links': [
            {'label': 'Sitio oficial Max Player', 'url': 'https://maxplayer.tv/en'},
        ],
        'app_options': [
            {
                'title': 'Opción 1 · Max Player',
                'subtitle': 'Disponible para Amazon Firestick, Android Stick y TV Box Android.',
                'logo': 'apps/max-player.webp',
                'steps': [
                    'Conecta tu dispositivo a internet.',
                    'Busca Max Player desde la tienda o método de instalación disponible.',
                    'Instala la app y ábrela.',
                    'Solicita tus datos de acceso por WhatsApp.',
                    'Ingresa usuario y contraseña y prueba un canal.'
                ],
                'links': [
                    {'label': 'Sitio oficial Max Player', 'url': 'https://maxplayer.tv/en'},
                ]
            },
            {
                'title': 'Opción 2 · XD Plus',
                'subtitle': 'Disponible para Amazon Firestick, Android Stick y TV Box Android.',
                'logo': 'apps/xd-plus.webp',
                'download_code': 'Downloader · código 8813174',
                'steps': [
                    'Descarga e instala la app Downloader en tu dispositivo.',
                    'Abre Downloader y escribe el código 8813174.',
                    'Instala XD Plus y abre la aplicación.',
                    'Escríbenos por WhatsApp para confirmar tus datos de acceso.'
                ],
                'links': []
            }
        ]
    },
    'samsung-lg': {
        'title': 'Samsung / LG Smart TV',
        'app': 'Max Player',
        'icon': '📺',
        'logo': 'apps/max-player.webp',
        'hero_image': 'landing/device-samsung-lg.webp',
        'steps': [
            'Abre la tienda de apps de tu Smart TV.',
            'Busca Max Player.',
            'Instala y abre la aplicación.',
            'Envíanos por WhatsApp el modelo de TV si necesitas ayuda.',
            'Ingresa los datos de acceso que te asignaremos.'
        ],
        'tips': ['Conecta la TV por cable de red si es posible.', 'Actualiza el software de la TV para mejor compatibilidad.'],
        'links': [
            {'label': 'Sitio oficial Max Player', 'url': 'https://maxplayer.tv/en'},
        ],
        'app_options': [
            {
                'title': 'Opción 1 · Max Player',
                'subtitle': 'Opción recomendada para Samsung y LG Smart TV.',
                'logo': 'apps/max-player.webp',
                'steps': [
                    'Abre la tienda de apps de tu Smart TV.',
                    'Busca Max Player.',
                    'Instala y abre la aplicación.',
                    'Envíanos por WhatsApp el modelo de TV si necesitas ayuda.',
                    'Ingresa los datos de acceso que te asignaremos.'
                ],
                'links': [
                    {'label': 'Sitio oficial Max Player', 'url': 'https://maxplayer.tv/en'},
                ]
            },
            {
                'title': 'Opción 2 · Smart One',
                'subtitle': 'La licencia de la app tiene un costo anual de 2.50 euros. Nosotros podemos pagarla, previo pago a Fénix.',
                'logo': 'apps/smart-one.webp',
                'steps': [
                    'Abre la tienda de apps de tu Smart TV.',
                    'Busca e instala Smart One IPTV.',
                    'Escríbenos por WhatsApp para coordinar la licencia anual de 2.50 euros.',
                    'Ventaja: permite agregar subtítulos cuando estén disponibles en la película o serie.',
                    'Ingresa o envíanos los datos que te solicite la app para configurarla.'
                ],
                'links': [
                    {'label': 'Sitio oficial Smart One', 'url': 'https://smartone-iptv.com/'},
                ]
            },
            {
                'title': 'Opción 3 · Ora Player',
                'subtitle': 'La licencia de la app tiene un costo anual de US$3. Nosotros podemos pagarla, previo pago a Fénix.',
                'logo': 'apps/ora-player.webp',
                'steps': [
                    'Abre la tienda de apps de tu Smart TV.',
                    'Busca e instala Ora Player.',
                    'Escríbenos por WhatsApp para coordinar la licencia anual de US$3.',
                    'Ventaja: permite agregar subtítulos o cambiar audio cuando esté disponible.',
                    'Ingresa o envíanos los datos que te solicite la app para configurarla.'
                ],
                'links': [
                    {'label': 'Sitio oficial Ora Player', 'url': 'https://oraplayerapps.com/'},
                ]
            }
        ]
    },
    'hisense': {
        'title': 'Hisense Smart TV',
        'app': 'Smart One',
        'icon': '🖥️',
        'logo': 'apps/smart-one.webp',
        'hero_image': 'landing/device-hisense.webp',
        'steps': [
            'Abre la tienda de aplicaciones de Hisense.',
            'Busca Smart One.',
            'Instala la app y ábrela.',
            'Envíanos la información que solicite la app si aplica.',
            'Te guiaremos por WhatsApp para cargar o activar el acceso.'
        ],
        'tips': ['Algunas versiones de Hisense pueden variar por país.', 'Si no aparece la app, escríbenos el modelo exacto.']
    },
    'roku': {
        'title': 'Roku TV',
        'app': 'XD+ / SACNET Plus / Premium Player',
        'icon': '🟣',
        'logo': 'apps/premium-player.webp',
        'alt_logo': 'apps/sacnet-plus.webp',
        'hero_image': 'landing/device-roku.webp',
        'steps': [
            'Agrega Premium Player o SACNET Plus en tu Roku.',
            'Abre la app y revisa la pantalla inicial.',
            'Escríbenos por WhatsApp para confirmar compatibilidad.',
            'Te enviaremos los datos de acceso o instrucciones necesarias.',
            'Prueba la reproducción durante la demo.'
        ],
        'tips': ['Roku puede variar según región/cuenta.', 'Si una app no aparece, te indicaremos alternativas disponibles.', 'Escríbenos el modelo de Roku si necesitas ayuda.'],
        'app_options': [
            {
                'title': 'Opción 1 · XD Plus',
                'subtitle': 'Primera opción para Roku, según disponibilidad por región/cuenta.',
                'logo': 'apps/xd-plus.webp',
                'steps': [
                    'Busca XD Plus en tu Roku o confirma con nosotros disponibilidad.',
                    'Instala la app y abre la pantalla inicial.',
                    'Escríbenos por WhatsApp para confirmar compatibilidad.',
                    'Te enviaremos los datos de acceso o instrucciones necesarias.',
                    'Prueba la reproducción durante la demo.'
                ],
                'links': []
            },
            {
                'title': 'Opción 2 · SACNET Plus',
                'subtitle': 'Alternativa para Roku cuando esté disponible en tu cuenta/región.',
                'logo': 'apps/sacnet-plus.webp',
                'steps': [
                    'Busca SACNET Plus en tu Roku.',
                    'Instala la app y abre la pantalla inicial.',
                    'Escríbenos por WhatsApp para confirmar compatibilidad.',
                    'Te enviaremos los datos de acceso o instrucciones necesarias.',
                    'Prueba la reproducción durante la demo.'
                ],
                'links': []
            },
            {
                'title': 'Opción 3 · Premium Player',
                'subtitle': 'Otra opción compatible para Roku, según disponibilidad.',
                'logo': 'apps/premium-player.webp',
                'steps': [
                    'Busca Premium Player en tu Roku.',
                    'Instala la app y abre la pantalla inicial.',
                    'Escríbenos por WhatsApp para confirmar compatibilidad.',
                    'Te enviaremos los datos de acceso o instrucciones necesarias.',
                    'Prueba la reproducción durante la demo.'
                ],
                'links': []
            }
        ]
    },
    'windows-mac': {
        'title': 'Windows / Mac',
        'app': 'Max Player',
        'icon': '💻',
        'logo': 'apps/max-player.webp',
        'hero_image': 'landing/device-windows-mac.webp',
        'steps': [
            'Instala Max Player o la app compatible que te indiquemos.',
            'Conecta tu computadora a una red estable.',
            'Solicita demo o activación por WhatsApp.',
            'Ingresa usuario y contraseña.',
            'Ajusta pantalla completa para mejor experiencia.'
        ],
        'tips': ['Cierra VPNs o extensiones que afecten la conexión.', 'Usa audífonos o salida HDMI si conectarás a TV.'],
        'links': [
            {'label': 'Max Player para Windows', 'url': 'https://apps.microsoft.com/detail/9NJP3PH1HXL6'},
            {'label': 'Sitio oficial Max Player', 'url': 'https://maxplayer.tv/en'},
        ]
    },
    'whatsapp': {
        'title': 'Atención por WhatsApp',
        'app': 'Soporte Fénix',
        'icon': '💬',
        'logo': '',
        'hero_image': 'landing/device-support.webp',
        'steps': [
            'Escríbenos al WhatsApp oficial.',
            'Indica país, dispositivo y si deseas demo o contratar.',
            'Te diremos qué app instalar.',
            'Validamos datos y pago cuando corresponda.',
            'Te acompañamos hasta que puedas ingresar.'
        ],
        'tips': ['Incluye captura si ves algún error.', 'Indica si usas WiFi, datos móviles o cable de red.']
    },
    'demo': {
        'title': 'Demo gratis de 3 horas',
        'app': 'Prueba Fénix',
        'icon': '⏱️',
        'logo': '',
        'hero_image': 'landing/device-demo.webp',
        'steps': [
            'Solicita la demo por WhatsApp.',
            'Indica tu dispositivo principal.',
            'Instala la app recomendada.',
            'Recibe tus datos temporales de prueba.',
            'Evalúa calidad, contenido y compatibilidad.'
        ],
        'tips': ['La demo es para validar compatibilidad.', 'La disponibilidad puede depender del horario de atención.']
    },
}


@app.route('/dispositivo/<slug>')
def device_guide(slug):
    guide = DEVICE_GUIDES.get(slug)
    if not guide:
        return redirect(url_for('public_home') + '#dispositivos')
    return render_template('device_guide.html', guide=guide, slug=slug)


@app.route('/cliente/login', methods=['GET', 'POST'])
def client_login():
    error = None
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()
        db = get_db()
        c = db.cursor()
        c.execute(qmark("""
            SELECT a.username, a.password, a.is_enabled, cl.nombre
            FROM client_portal_accounts a
            LEFT JOIN clientes cl ON cl.username = a.username
            WHERE a.username=?
        """), (username,))
        account = fetchone(c)
        if not account:
            c.execute(qmark("""
                SELECT a.username, a.password, a.is_enabled, cl.nombre
                FROM client_portal_accounts a
                LEFT JOIN clientes cl ON cl.username = a.username
                WHERE LOWER(a.username)=LOWER(?)
            """), (username,))
            account = fetchone(c)
        if account and account.get('is_enabled') and verify_password(account.get('password'), password):
            session.clear()
            session['client_username'] = account['username']
            session['client_name'] = account.get('nombre') or account['username']
            db.close()
            return redirect(url_for('client_portal'))
        c.execute(qmark("SELECT username, nombre FROM clientes WHERE username=?"), (username,))
        client = fetchone(c)
        if not client:
            c.execute(qmark("SELECT username, nombre FROM clientes WHERE LOWER(username)=LOWER(?)"), (username,))
            client = fetchone(c)
        if client and password == client['username']:
            now = now_gt().isoformat(timespec='seconds')
            if PG:
                c.execute("""
                    INSERT INTO client_portal_accounts (username, password, is_enabled, updated_at)
                    VALUES (%s,%s,TRUE,%s)
                    ON CONFLICT (username) DO UPDATE SET
                        password=EXCLUDED.password,
                        is_enabled=TRUE,
                        updated_at=EXCLUDED.updated_at
                """, (client['username'], hash_password(client['username']), now))
            else:
                c.execute("""
                    INSERT INTO client_portal_accounts (username, password, is_enabled, updated_at)
                    VALUES (?,?,1,?)
                    ON CONFLICT(username) DO UPDATE SET
                        password=excluded.password,
                        is_enabled=1,
                        updated_at=excluded.updated_at
                """, (client['username'], hash_password(client['username']), now))
            db.commit()
            session.clear()
            session['client_username'] = client['username']
            session['client_name'] = client.get('nombre') or client['username']
            db.close()
            return redirect(url_for('client_portal'))
        db.close()
        error = 'Usuario o contraseña incorrectos, o acceso no habilitado.'
    return render_template('client_login.html', error=error)


@app.route('/cliente/logout')
def client_logout():
    session.pop('client_username', None)
    session.pop('client_name', None)
    return redirect(url_for('client_login'))


@app.route('/cliente')
@client_login_required
def client_portal():
    username = session['client_username']
    db = get_db()
    c = db.cursor()
    c.execute(qmark("SELECT * FROM clientes WHERE username=?"), (username,))
    client = fetchone(c) or {'username': username, 'nombre': username}
    c.execute(qmark("SELECT * FROM client_service_credentials WHERE username=?"), (username,))
    service = fetchone(c)
    if not service:
        c.execute(qmark("""
            SELECT username, xui_username as service_username, xui_password as service_password,
                   xui_expires_at as expires_at, completed_at
            FROM activation_tasks
            WHERE username=? AND status='done'
            ORDER BY completed_at DESC, id DESC
            LIMIT 1
        """), (username,))
        latest = fetchone(c)
        if latest:
            service = {
                'app_name': 'Max Player',
                'service_username': latest.get('service_username'),
                'service_password': latest.get('service_password'),
                'expires_at': latest.get('expires_at'),
                'devices': 3,
            }
    c.execute(qmark("SELECT mes, monto, fecha_registro FROM pagos WHERE username=? ORDER BY mes DESC LIMIT 6"), (username,))
    payments = fetchall(c)
    parent = None
    asociados = []
    parent_username = (client or {}).get('parent_username')
    if parent_username:
        c.execute(qmark("SELECT username, nombre, contacto, vencimiento FROM clientes WHERE username=?"), (parent_username,))
        parent = fetchone(c)
        c.execute(qmark("""
            SELECT username, nombre, contacto, vencimiento
            FROM clientes
            WHERE parent_username=? AND username<>?
            ORDER BY nombre, username
        """), (parent_username, username))
        asociados = fetchall(c)
    else:
        c.execute(qmark("""
            SELECT username, nombre, contacto, vencimiento
            FROM clientes
            WHERE parent_username=?
            ORDER BY nombre, username
        """), (username,))
        asociados = fetchall(c)
    db.close()
    expires_at = (service or {}).get('expires_at') or client.get('vencimiento')
    status = service_status(expires_at)
    return render_template('client_portal.html',
                           client=client,
                           service=service or {},
                           payments=payments,
                           parent=parent,
                           asociados=asociados,
                           expires_at=expires_at,
                           status=status)


@app.route('/api/cliente/cambiar-password', methods=['POST'])
@client_login_required
def client_change_password():
    data = request.json or {}
    current_password = data.get('password_actual', '')
    new_password = data.get('password_nueva', '')
    confirm_password = data.get('password_confirmar', '')
    if not current_password or not new_password or len(new_password) < 6 or new_password != confirm_password:
        return jsonify({'error': 'Datos inválidos. La nueva contraseña debe tener al menos 6 caracteres y coincidir con la confirmación.'}), 400
    username = session['client_username']
    db = get_db()
    c = db.cursor()
    c.execute(qmark("SELECT password FROM client_portal_accounts WHERE username=? AND is_enabled=?"),
              (username, True if PG else 1))
    account = fetchone(c)
    if not account or not verify_password(account.get('password'), current_password):
        db.close()
        return jsonify({'error': 'La contraseña actual es incorrecta.'}), 403
    now = now_gt().isoformat(timespec='seconds')
    c.execute(qmark("UPDATE client_portal_accounts SET password=?, updated_at=? WHERE username=?"),
              (hash_password(new_password), now, username))
    db.commit()
    db.close()
    return jsonify({'ok': True})


@app.route('/api/cliente/actualizar-datos', methods=['POST'])
@client_login_required
def client_update_profile():
    data = request.json or {}
    username = session['client_username']
    contacto = ''.join(ch for ch in (data.get('contacto') or '') if ch.isdigit())[-15:]
    contacto_secundario = ''.join(ch for ch in (data.get('contacto_secundario') or '') if ch.isdigit())[-15:]
    email = (data.get('email') or '').strip().lower()
    if not contacto or len(contacto) < 8:
        return jsonify({'error': 'Ingresa un celular principal válido.'}), 400
    if email and ('@' not in email or '.' not in email.split('@')[-1]):
        return jsonify({'error': 'Ingresa un correo electrónico válido.'}), 400
    now = now_gt().isoformat(timespec='seconds')
    db = get_db()
    c = db.cursor()
    c.execute(qmark("""
        UPDATE clientes
        SET contacto=?, contacto_secundario=?, email=?, datos_actualizados_at=?
        WHERE username=?
    """), (contacto, contacto_secundario, email, now, username))
    db.commit()
    db.close()
    return jsonify({'ok': True})


@app.route('/login', methods=['GET', 'POST'])
def login():
    error = None
    if request.method == 'POST':
        u = request.form.get('username', '').strip()
        p = request.form.get('password', '').strip()
        db = get_db()
        c = db.cursor()
        c.execute(qmark("SELECT * FROM usuarios WHERE username=?"), (u,))
        user = fetchone(c)
        if user and verify_password(user.get('password'), p):
            maybe_upgrade_password_hash(c, u, user.get('password'), p)
            db.commit()
            db.close()
            session['user'] = u
            session['rol'] = user['rol']
            return redirect(url_for('index'))
        db.close()
        error = 'Usuario o contraseña incorrectos'
    return render_template('login.html', error=error)

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

@app.route('/app')
@login_required
def index():
    return render_template('index.html', user=session['user'], rol=session['rol'])

# ── API: DASHBOARD ────────────────────────────────────────────────────────────
@app.route('/api/dashboard')
@login_required
def dashboard():
    db = get_db()
    c = db.cursor()
    today = today_gt().isoformat()
    in_30 = today_gt().replace(day=min(today_gt().day + 30, 28)).isoformat()
    mes_actual = today_gt().strftime('%Y-%m')

    c.execute(qmark("SELECT COUNT(*) as c FROM clientes WHERE vencimiento >= ?"), (today,))
    activos = fetchone(c)['c']

    c.execute(qmark("SELECT COUNT(*) as c FROM clientes WHERE vencimiento < ? OR vencimiento IS NULL"), (today,))
    vencidos = fetchone(c)['c']

    c.execute(qmark("SELECT COUNT(*) as c FROM clientes WHERE vencimiento >= ? AND vencimiento <= ?"), (today, in_30))
    por_vencer = fetchone(c)['c']

    c.execute(qmark("SELECT COALESCE(SUM(monto),0) as t FROM pagos WHERE mes LIKE ?"), (f'{mes_actual}%',))
    ingresos = fetchone(c)['t']

    t = today_gt()
    m6 = t.month - 6
    y6 = t.year + (m6 - 1) // 12
    m6 = ((m6 - 1) % 12) + 1
    six_ago_str = f"{y6}-{m6:02d}"
    if PG:
        c.execute("""
            SELECT TO_CHAR(mes::date, 'YYYY-MM') as m,
                   SUM(monto) as total,
                   COUNT(DISTINCT username) as clientes,
                   COUNT(*) as unidades
            FROM pagos WHERE SUBSTRING(mes,1,7) >= %s
            GROUP BY m ORDER BY m
        """, (six_ago_str,))
    else:
        c.execute("""
            SELECT strftime('%Y-%m', mes) as m,
                   SUM(monto) as total,
                   COUNT(DISTINCT username) as clientes,
                   COUNT(*) as unidades
            FROM pagos WHERE substr(mes,1,7) >= ?
            GROUP BY m ORDER BY m
        """, (six_ago_str,))
    ultimos_meses = fetchall(c)

    c.execute(qmark("""
        SELECT username, nombre, contacto, vencimiento FROM clientes
        WHERE vencimiento >= ? AND vencimiento <= ?
        ORDER BY vencimiento LIMIT 20
    """), (today, in_30))
    pronto = fetchall(c)
    db.close()

    return jsonify({
        'activos': activos, 'vencidos': vencidos, 'por_vencer': por_vencer,
        'ingresos_mes': round(float(ingresos), 2),
        'ultimos_meses': ultimos_meses,
        'por_vencer_lista': pronto
    })

# ── API: SYNC (para reporte_diario) ──────────────────────────────────────────
@app.route('/api/sync/clientes')
@login_required
def sync_clientes():
    """Devuelve todos los usernames + vencimiento de una sola vez (sin paginación).
    Solo para uso interno del reporte_diario.py."""
    db = get_db()
    c = db.cursor()
    c.execute("SELECT username, vencimiento FROM clientes ORDER BY username ASC")
    rows = fetchall(c)
    db.close()
    return jsonify({'clientes': rows, 'total': len(rows)})

# ── API: CLIENTES ─────────────────────────────────────────────────────────────
@app.route('/api/clientes')
@login_required
def clientes():
    q = request.args.get('q', '').strip()
    estado = request.args.get('estado', '')
    fecha = request.args.get('fecha', '').strip()
    page = int(request.args.get('page', 1))
    per_page = 20
    offset = (page - 1) * per_page
    today = today_gt().isoformat()

    where, params = [], []
    if q:
        where.append("(nombre ILIKE ? OR username ILIKE ? OR contacto ILIKE ?)" if PG else
                     "(nombre LIKE ? OR username LIKE ? OR contacto LIKE ?)")
        params += [f'%{q}%', f'%{q}%', f'%{q}%']
    if fecha:
        where.append("vencimiento = ?"); params.append(fecha)
    elif estado == 'activo':
        where.append("vencimiento >= ?"); params.append(today)
    elif estado == 'vencido':
        where.append("(vencimiento < ? OR vencimiento IS NULL)"); params.append(today)
    elif estado == 'por_vencer':
        in_30 = today_gt().replace(day=min(today_gt().day + 30, 28)).isoformat()
        where.append("vencimiento >= ? AND vencimiento <= ?"); params += [today, in_30]

    where_sql = "WHERE " + " AND ".join(where) if where else ""
    db = get_db()
    c = db.cursor()
    c.execute(qmark(f"SELECT COUNT(*) as c FROM clientes {where_sql}"), params)
    total = fetchone(c)['c']
    c.execute(qmark(f"""
        SELECT username, nombre, contacto, vencimiento, referido, parent_username, total_pagado, notas
        FROM clientes {where_sql} ORDER BY vencimiento DESC, username ASC LIMIT ? OFFSET ?
    """), params + [per_page, offset])
    rows = fetchall(c)
    db.close()
    return jsonify({'total': total, 'clientes': rows})

@app.route('/api/clientes', methods=['POST'])
@login_required
def crear_cliente():
    data = request.json
    username = data.get('username', '').strip()
    if not username:
        return jsonify({'error': 'El username es obligatorio'}), 400
    db = get_db()
    c = db.cursor()
    c.execute(qmark("SELECT username FROM clientes WHERE username=?"), (username,))
    if fetchone(c):
        db.close()
        return jsonify({'error': 'Ese username ya existe'}), 409
    parent_username = (data.get('parent_username') or '').strip() or None
    if parent_username:
        if parent_username == username:
            db.close()
            return jsonify({'error': 'Un cliente no puede ser su propio titular'}), 400
        c.execute(qmark("SELECT username FROM clientes WHERE username=?"), (parent_username,))
        if not fetchone(c):
            db.close()
            return jsonify({'error': 'El usuario titular/principal no existe'}), 400
    c.execute(qmark("""
        INSERT INTO clientes (username, nombre, contacto, vencimiento, referido, parent_username, notas, total_pagado)
        VALUES (?,?,?,?,?,?,?,0)
    """), (username, data.get('nombre',''), data.get('contacto',''),
           data.get('vencimiento') or None, data.get('referido','NO'), parent_username, data.get('notas','')))
    db.commit()
    db.close()
    return jsonify({'ok': True})

@app.route('/api/clientes/<username>')
@login_required
def cliente_detalle(username):
    db = get_db()
    c = db.cursor()
    c.execute(qmark("SELECT * FROM clientes WHERE username=?"), (username,))
    cliente = fetchone(c)
    if not cliente:
        db.close()
        return jsonify({'error': 'No encontrado'}), 404

    rol = session.get('rol', 'atencion')
    # Ambos roles ven el historial de pagos reciente
    c.execute(qmark("SELECT id, mes, monto, fecha_registro, (comprobante IS NOT NULL) as has_comprobante FROM pagos WHERE username=? ORDER BY mes DESC LIMIT 24"), (username,))
    pagos = fetchall(c)
    c.execute(qmark("""
        SELECT service_username, service_password, expires_at,
               maxplayer_user_id, maxplayer_synced_at, maxplayer_sync_status
        FROM client_service_credentials WHERE username=?
    """), (username,))
    service = fetchone(c) or {}
    parent = None
    if cliente.get('parent_username'):
        c.execute(qmark("""
            SELECT username, nombre, contacto, vencimiento, total_pagado
            FROM clientes WHERE username=?
        """), (cliente.get('parent_username'),))
        parent = fetchone(c)
    c.execute(qmark("""
        SELECT username, nombre, contacto, vencimiento, total_pagado
        FROM clientes
        WHERE parent_username=?
        ORDER BY nombre ASC, username ASC
    """), (username,))
    asociados = fetchall(c)
    db.close()
    return jsonify({'cliente': cliente, 'pagos': pagos, 'rol': rol, 'service': service, 'parent': parent, 'asociados': asociados})


@app.route('/api/clientes/<username>/portal', methods=['POST'])
@login_required
def actualizar_portal_cliente(username):
    data = request.json or {}
    portal_password = (data.get('portal_password') or '').strip()
    enabled = bool(data.get('enabled', True))
    service_username = (data.get('service_username') or '').strip()
    service_password = (data.get('service_password') or '').strip()
    app_name = (data.get('app_name') or 'Max Player').strip()
    expires_at = (data.get('expires_at') or '').strip()
    devices = int(data.get('devices') or 3)
    now = now_gt().isoformat(timespec='seconds')

    db = get_db()
    c = db.cursor()
    c.execute(qmark("SELECT username FROM clientes WHERE username=?"), (username,))
    if not fetchone(c):
        db.close()
        return jsonify({'error': 'Cliente no encontrado'}), 404

    if portal_password:
        if PG:
            c.execute("""
                INSERT INTO client_portal_accounts (username, password, is_enabled, updated_at)
                VALUES (%s,%s,%s,%s)
                ON CONFLICT (username) DO UPDATE SET
                    password=EXCLUDED.password,
                    is_enabled=EXCLUDED.is_enabled,
                    updated_at=EXCLUDED.updated_at
            """, (username, hash_password(portal_password), enabled, now))
        else:
            c.execute("""
                INSERT INTO client_portal_accounts (username, password, is_enabled, updated_at)
                VALUES (?,?,?,?)
                ON CONFLICT(username) DO UPDATE SET
                    password=excluded.password,
                    is_enabled=excluded.is_enabled,
                    updated_at=excluded.updated_at
            """, (username, hash_password(portal_password), 1 if enabled else 0, now))
    else:
        c.execute(qmark("UPDATE client_portal_accounts SET is_enabled=?, updated_at=? WHERE username=?"),
                  (enabled, now, username))

    if service_username or service_password or expires_at:
        if PG:
            c.execute("""
                INSERT INTO client_service_credentials
                    (username, app_name, service_username, service_password, expires_at, devices, updated_at)
                VALUES (%s,%s,%s,%s,%s,%s,%s)
                ON CONFLICT (username) DO UPDATE SET
                    app_name=EXCLUDED.app_name,
                    service_username=COALESCE(NULLIF(EXCLUDED.service_username,''), client_service_credentials.service_username),
                    service_password=COALESCE(NULLIF(EXCLUDED.service_password,''), client_service_credentials.service_password),
                    expires_at=COALESCE(NULLIF(EXCLUDED.expires_at,''), client_service_credentials.expires_at),
                    devices=EXCLUDED.devices,
                    updated_at=EXCLUDED.updated_at
            """, (username, app_name, service_username, service_password, expires_at, devices, now))
        else:
            c.execute("""
                INSERT INTO client_service_credentials
                    (username, app_name, service_username, service_password, expires_at, devices, updated_at)
                VALUES (?,?,?,?,?,?,?)
                ON CONFLICT(username) DO UPDATE SET
                    app_name=excluded.app_name,
                    service_username=COALESCE(NULLIF(excluded.service_username,''), client_service_credentials.service_username),
                    service_password=COALESCE(NULLIF(excluded.service_password,''), client_service_credentials.service_password),
                    expires_at=COALESCE(NULLIF(excluded.expires_at,''), client_service_credentials.expires_at),
                    devices=excluded.devices,
                    updated_at=excluded.updated_at
            """, (username, app_name, service_username, service_password, expires_at, devices, now))

    db.commit()
    db.close()
    return jsonify({'ok': True})

@app.route('/api/clientes/<username>/portal/reset-password', methods=['POST'])
@login_required
def restablecer_password_portal_cliente(username):
    db = get_db()
    c = db.cursor()
    c.execute(qmark("SELECT username FROM clientes WHERE username=?"), (username,))
    cliente = fetchone(c)
    if not cliente:
        db.close()
        return jsonify({'error': 'Cliente no encontrado'}), 404
    now = now_gt().isoformat(timespec='seconds')
    if PG:
        c.execute("""
            INSERT INTO client_portal_accounts (username, password, is_enabled, updated_at)
            VALUES (%s,%s,TRUE,%s)
            ON CONFLICT (username) DO UPDATE SET
                password=EXCLUDED.password,
                is_enabled=TRUE,
                updated_at=EXCLUDED.updated_at
        """, (username, hash_password(username), now))
    else:
        c.execute("""
            INSERT INTO client_portal_accounts (username, password, is_enabled, updated_at)
            VALUES (?,?,1,?)
            ON CONFLICT(username) DO UPDATE SET
                password=excluded.password,
                is_enabled=1,
                updated_at=excluded.updated_at
        """, (username, hash_password(username), now))
    db.commit()
    db.close()
    return jsonify({'ok': True, 'username': username, 'password': username})

@app.route('/api/clientes/<username>/maxplayer/restore', methods=['POST'])
@login_required
def restaurar_maxplayer_cliente(username):
    data = request.json or {}
    provided_service_password = (data.get('service_password') or '').strip()
    db = get_db()
    c = db.cursor()
    c.execute(qmark("SELECT nombre, vencimiento FROM clientes WHERE username=?"), (username,))
    cliente = fetchone(c)
    if not cliente:
        db.close()
        return jsonify({'error': 'Cliente no encontrado'}), 404
    c.execute(qmark("""
        SELECT service_username, service_password, expires_at, maxplayer_user_id
        FROM client_service_credentials WHERE username=?
    """), (username,))
    service = fetchone(c)
    if not service:
        service = {
            'service_username': username,
            'service_password': '',
            'expires_at': cliente.get('vencimiento'),
            'maxplayer_user_id': None,
        }
    if not service.get('service_username'):
        service['service_username'] = username
    if provided_service_password:
        service['service_password'] = provided_service_password
    if not service.get('service_password'):
        db.close()
        return jsonify({'error': 'Este cliente aun no tiene password XUI/IPTV guardado. Ingresa el password XUI para restaurar Max Player.'}), 400

    service_username = service['service_username']
    service_password = service['service_password']
    try:
        existing_user_id = service.get('maxplayer_user_id') or find_maxplayer_user_id(service_username)
        if existing_user_id:
            purge_maxplayer_user(existing_user_id)
        response, maxplayer_user_id = create_maxplayer_user(
            username=service_username,
            iptv_user=service_username,
            iptv_pass=service_password,
            password=service_password,
            fullname=cliente.get('nombre') or username,
            user_email=''
        )
    except MaxPlayerError as exc:
        db.close()
        return jsonify({'error': str(exc)}), 400

    now = datetime.now(GT_TZ).isoformat()
    sync_status = 'restored' if maxplayer_user_id else 'restored_no_id'
    if PG:
        c.execute("""
            INSERT INTO client_service_credentials
                (username, app_name, service_username, service_password, expires_at,
                 maxplayer_user_id, maxplayer_synced_at, maxplayer_sync_status, updated_at)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
            ON CONFLICT (username) DO UPDATE SET
                app_name=EXCLUDED.app_name,
                service_username=EXCLUDED.service_username,
                service_password=EXCLUDED.service_password,
                expires_at=COALESCE(EXCLUDED.expires_at, client_service_credentials.expires_at),
                maxplayer_user_id=EXCLUDED.maxplayer_user_id,
                maxplayer_synced_at=EXCLUDED.maxplayer_synced_at,
                maxplayer_sync_status=EXCLUDED.maxplayer_sync_status,
                updated_at=EXCLUDED.updated_at
        """, (username, 'Max Player', service_username, service_password, service.get('expires_at'),
              maxplayer_user_id, now, sync_status, now))
    else:
        c.execute("""
            INSERT INTO client_service_credentials
                (username, app_name, service_username, service_password, expires_at,
                 maxplayer_user_id, maxplayer_synced_at, maxplayer_sync_status, updated_at)
            VALUES (?,?,?,?,?,?,?,?,?)
            ON CONFLICT(username) DO UPDATE SET
                app_name=excluded.app_name,
                service_username=excluded.service_username,
                service_password=excluded.service_password,
                expires_at=COALESCE(excluded.expires_at, client_service_credentials.expires_at),
                maxplayer_user_id=excluded.maxplayer_user_id,
                maxplayer_synced_at=excluded.maxplayer_synced_at,
                maxplayer_sync_status=excluded.maxplayer_sync_status,
                updated_at=excluded.updated_at
        """, (username, 'Max Player', service_username, service_password, service.get('expires_at'),
              maxplayer_user_id, now, sync_status, now))
    db.commit()
    db.close()
    return jsonify({'ok': True, 'maxplayer_user_id': maxplayer_user_id, 'status': sync_status})

@app.route('/api/clientes/<username>', methods=['PUT'])
@login_required
def actualizar_cliente(username):
    data = request.json or {}
    db = get_db()
    c = db.cursor()
    c.execute(qmark("SELECT username FROM clientes WHERE username=?"), (username,))
    if not fetchone(c):
        db.close()
        return jsonify({'error': 'Cliente no encontrado'}), 404

    # Cambio de username (solo admin)
    nuevo_username = data.get('nuevo_username', '').strip()
    if nuevo_username and nuevo_username != username:
        if session.get('rol') != 'admin':
            db.close()
            return jsonify({'error': 'Sin permiso para cambiar username'}), 403
        # Verificar que el nuevo username no exista
        c.execute(qmark("SELECT username FROM clientes WHERE username=?"), (nuevo_username,))
        if fetchone(c):
            db.close()
            return jsonify({'error': 'Ese username ya está en uso'}), 409
        related_tables = [
            ('client_portal_accounts', 'portal del cliente'),
            ('client_service_credentials', 'credenciales del servicio'),
        ]
        for table, label in related_tables:
            c.execute(qmark(f"SELECT username FROM {table} WHERE username=?"), (nuevo_username,))
            if fetchone(c):
                db.close()
                return jsonify({'error': f'El nuevo username ya existe en {label}. Revisa si ese usuario fue creado antes.'}), 409
        # Actualizar pagos primero (FK)
        c.execute(qmark("UPDATE pagos SET username=? WHERE username=?"), (nuevo_username, username))
        c.execute(qmark("UPDATE client_portal_accounts SET username=? WHERE username=?"), (nuevo_username, username))
        c.execute(qmark("UPDATE client_service_credentials SET username=? WHERE username=?"), (nuevo_username, username))
        c.execute(qmark("UPDATE activation_tasks SET username=? WHERE username=?"), (nuevo_username, username))
        c.execute(qmark("UPDATE clientes SET parent_username=? WHERE parent_username=?"), (nuevo_username, username))
        # Actualizar cliente
        c.execute(qmark("UPDATE clientes SET username=? WHERE username=?"), (nuevo_username, username))
        username = nuevo_username  # usar el nuevo para el resto de campos

    fields, params = [], []
    for field in ['nombre', 'contacto', 'vencimiento', 'referido', 'notas', 'parent_username']:
        if field in data:
            value = data[field]
            if field == 'parent_username':
                value = (value or '').strip() or None
                if value == username:
                    db.close()
                    return jsonify({'error': 'Un cliente no puede ser su propio titular'}), 400
                if value:
                    c.execute(qmark("SELECT username FROM clientes WHERE username=?"), (value,))
                    if not fetchone(c):
                        db.close()
                        return jsonify({'error': 'El usuario titular/principal no existe'}), 400
                    c.execute(qmark("SELECT parent_username FROM clientes WHERE username=?"), (value,))
                    titular = fetchone(c)
                    if titular and titular.get('parent_username') == username:
                        db.close()
                        return jsonify({'error': 'No se puede crear una relación circular entre clientes'}), 400
            fields.append(f"{field}=?")
            params.append(value)
    if fields:
        params.append(username)
        c.execute(qmark(f"UPDATE clientes SET {', '.join(fields)} WHERE username=?"), params)

    service_username = (data.get('service_username') or '').strip()
    service_password = (data.get('service_password') or '').strip()
    if service_username or service_password or data.get('vencimiento'):
        now = now_gt().isoformat(timespec='seconds')
        expires_at = data.get('vencimiento') or ''
        if PG:
            c.execute("""
                INSERT INTO client_service_credentials
                    (username, app_name, service_username, service_password, expires_at, devices, updated_at)
                VALUES (%s,%s,%s,%s,%s,%s,%s)
                ON CONFLICT (username) DO UPDATE SET
                    app_name=EXCLUDED.app_name,
                    service_username=COALESCE(NULLIF(EXCLUDED.service_username,''), client_service_credentials.service_username),
                    service_password=COALESCE(NULLIF(EXCLUDED.service_password,''), client_service_credentials.service_password),
                    expires_at=COALESCE(NULLIF(EXCLUDED.expires_at,''), client_service_credentials.expires_at),
                    devices=EXCLUDED.devices,
                    maxplayer_user_id=CASE
                        WHEN NULLIF(EXCLUDED.service_username,'') IS NOT NULL
                             AND EXCLUDED.service_username <> client_service_credentials.service_username
                        THEN NULL
                        ELSE client_service_credentials.maxplayer_user_id
                    END,
                    maxplayer_sync_status=CASE
                        WHEN NULLIF(EXCLUDED.service_username,'') IS NOT NULL
                             AND EXCLUDED.service_username <> client_service_credentials.service_username
                        THEN 'pending_restore'
                        ELSE client_service_credentials.maxplayer_sync_status
                    END,
                    updated_at=EXCLUDED.updated_at
            """, (username, 'Max Player', service_username, service_password, expires_at, 3, now))
        else:
            c.execute("""
                INSERT INTO client_service_credentials
                    (username, app_name, service_username, service_password, expires_at, devices, updated_at)
                VALUES (?,?,?,?,?,?,?)
                ON CONFLICT(username) DO UPDATE SET
                    app_name=excluded.app_name,
                    service_username=COALESCE(NULLIF(excluded.service_username,''), client_service_credentials.service_username),
                    service_password=COALESCE(NULLIF(excluded.service_password,''), client_service_credentials.service_password),
                    expires_at=COALESCE(NULLIF(excluded.expires_at,''), client_service_credentials.expires_at),
                    devices=excluded.devices,
                    maxplayer_user_id=CASE
                        WHEN NULLIF(excluded.service_username,'') IS NOT NULL
                             AND excluded.service_username <> client_service_credentials.service_username
                        THEN NULL
                        ELSE client_service_credentials.maxplayer_user_id
                    END,
                    maxplayer_sync_status=CASE
                        WHEN NULLIF(excluded.service_username,'') IS NOT NULL
                             AND excluded.service_username <> client_service_credentials.service_username
                        THEN 'pending_restore'
                        ELSE client_service_credentials.maxplayer_sync_status
                    END,
                    updated_at=excluded.updated_at
            """, (username, 'Max Player', service_username, service_password, expires_at, 3, now))

    db.commit()
    db.close()
    return jsonify({'ok': True, 'nuevo_username': username})

@app.route('/api/clientes/<username>/auto-asociar', methods=['POST'])
@login_required
def auto_asociar_clientes(username):
    db = get_db()
    c = db.cursor()
    c.execute(qmark("SELECT username, nombre, contacto FROM clientes WHERE username=?"), (username,))
    titular = fetchone(c)
    if not titular:
        db.close()
        return jsonify({'error': 'Cliente titular no encontrado'}), 404
    titular_phone = normalize_phone(titular.get('contacto'))
    if len(titular_phone) < 8:
        db.close()
        return jsonify({'error': 'El cliente titular no tiene un teléfono válido para buscar coincidencias.'}), 400

    c.execute(qmark("""
        SELECT username, nombre, contacto, parent_username
        FROM clientes
        WHERE username<>?
          AND (parent_username IS NULL OR parent_username='')
    """), (username,))
    candidatos = fetchall(c)
    matches = [row for row in candidatos if normalize_phone(row.get('contacto')) == titular_phone]
    for row in matches:
        c.execute(qmark("UPDATE clientes SET parent_username=? WHERE username=?"), (username, row.get('username')))
    db.commit()
    db.close()
    return jsonify({
        'ok': True,
        'titular': username,
        'criterio': 'telefono',
        'asociados': [{'username': r.get('username'), 'nombre': r.get('nombre'), 'contacto': r.get('contacto')} for r in matches],
        'count': len(matches),
    })

@app.route('/api/clientes/grupos-telefono')
@login_required
def grupos_por_telefono():
    today = today_gt().isoformat()
    db = get_db()
    c = db.cursor()
    c.execute(qmark("""
        SELECT username, nombre, contacto, vencimiento, parent_username, total_pagado
        FROM clientes
        WHERE contacto IS NOT NULL
          AND contacto <> ''
          AND vencimiento IS NOT NULL
          AND vencimiento <> ''
          AND vencimiento >= ?
        ORDER BY contacto ASC, nombre ASC, username ASC
    """), (today,))
    rows = fetchall(c)
    db.close()

    grouped = {}
    for row in rows:
        phone = normalize_phone(row.get('contacto'))
        if len(phone) < 8:
            continue
        grouped.setdefault(phone, []).append(row)

    groups = []
    for phone, clientes in grouped.items():
        if len(clientes) < 2:
            continue
        sin_titular = [c for c in clientes if not c.get('parent_username')]
        sugerido = max(clientes, key=lambda c: float(c.get('total_pagado') or 0))
        groups.append({
            'telefono': phone,
            'pais': phone_country_label(phone),
            'total': len(clientes),
            'sin_titular': len(sin_titular),
            'titular_sugerido': sugerido.get('username'),
            'clientes': [{
                'username': c.get('username'),
                'nombre': c.get('nombre'),
                'contacto': c.get('contacto'),
                'telefono_key': normalize_phone(c.get('contacto')),
                'pais': phone_country_label(c.get('contacto')),
                'vencimiento': c.get('vencimiento'),
                'parent_username': c.get('parent_username') or '',
                'total_pagado': c.get('total_pagado') or 0,
            } for c in clientes],
        })
    groups.sort(key=lambda g: (g['sin_titular'], g['total']), reverse=True)
    return jsonify({'ok': True, 'groups': groups, 'total': len(groups)})

@app.route('/api/clientes/duplicados')
@login_required
def clientes_duplicados():
    """Detecta posibles duplicados usando únicamente clientes guardados en la plataforma."""
    today = today_gt().isoformat()
    solo_activos = request.args.get('activos', '1') != '0'
    db = get_db()
    c = db.cursor()
    if solo_activos:
        c.execute(qmark("""
            SELECT username, nombre, contacto, vencimiento, parent_username, total_pagado
            FROM clientes
            WHERE vencimiento IS NOT NULL
              AND vencimiento <> ''
              AND vencimiento >= ?
            ORDER BY nombre ASC, username ASC
        """), (today,))
    else:
        c.execute("""
            SELECT username, nombre, contacto, vencimiento, parent_username, total_pagado
            FROM clientes
            ORDER BY nombre ASC, username ASC
        """)
    rows = fetchall(c)
    db.close()

    def client_payload(row):
        return {
            'username': row.get('username') or '',
            'nombre': row.get('nombre') or '',
            'contacto': row.get('contacto') or '',
            'telefono_key': normalize_phone(row.get('contacto')),
            'pais': phone_country_label(row.get('contacto')),
            'vencimiento': row.get('vencimiento') or '',
            'parent_username': row.get('parent_username') or '',
            'total_pagado': row.get('total_pagado') or 0,
        }

    by_name = {}
    by_username = {}
    for row in rows:
        username = row.get('username') or ''
        name_key = normalize_text_key(row.get('nombre'))
        if name_key and len(name_key) >= 4:
            by_name.setdefault(name_key, []).append(row)

        username_key = normalize_username_base(username)
        if username_key and len(username_key) >= 5:
            by_username.setdefault(username_key, []).append(row)

    def build_groups(grouped):
        out = []
        for key, items in grouped.items():
            unique_usernames = {str(i.get('username') or '').lower() for i in items}
            if len(items) < 2 or len(unique_usernames) < 2:
                continue
            out.append({
                'key': key,
                'total': len(items),
                'clientes': [client_payload(i) for i in items],
            })
        out.sort(key=lambda g: (-g['total'], g['key']))
        return out[:150]

    name_groups = build_groups(by_name)
    username_groups = build_groups(by_username)

    return jsonify({
        'ok': True,
        'solo_activos': solo_activos,
        'counts': {
            'clientes_revisados': len(rows),
            'nombres_repetidos': len(name_groups),
            'usuarios_similares': len(username_groups),
        },
        'name_groups': name_groups,
        'username_groups': username_groups,
    })

@app.route('/api/clientes/<username>', methods=['DELETE'])
@login_required
def eliminar_cliente(username):
    if session.get('rol') != 'admin':
        return jsonify({'error': 'Acceso denegado'}), 403
    db = get_db()
    c = db.cursor()
    c.execute(qmark("SELECT username FROM clientes WHERE username=?"), (username,))
    if not fetchone(c):
        db.close()
        return jsonify({'error': 'Cliente no encontrado'}), 404
    c.execute(qmark("UPDATE clientes SET parent_username=NULL WHERE parent_username=?"), (username,))
    # Eliminar pagos asociados primero
    c.execute(qmark("DELETE FROM pagos WHERE username=?"), (username,))
    # Luego eliminar el cliente
    c.execute(qmark("DELETE FROM clientes WHERE username=?"), (username,))
    db.commit()
    db.close()
    return jsonify({'ok': True})

# ── API: PAGOS ────────────────────────────────────────────────────────────────
@app.route('/api/pagos', methods=['POST'])
@login_required
def registrar_pago():
    data = request.json or {}
    username = (data.get('username') or '').strip()
    try:
        monto = float(data.get('monto', 0))
    except (TypeError, ValueError):
        monto = 0
    vencimiento_nuevo = data.get('vencimiento')
    mes = data.get('mes', today_gt().strftime('%Y-%m-01'))
    comprobante = data.get('comprobante')  # base64 data URL, optional
    if not username or monto <= 0:
        return jsonify({'error': 'Datos inválidos'}), 400
    db = get_db()
    c = db.cursor()
    c.execute(qmark("SELECT username FROM clientes WHERE username=?"), (username,))
    if not fetchone(c):
        db.close()
        return jsonify({'error': f'Cliente no encontrado: {username}'}), 404
    if PG:
        c.execute("""
            INSERT INTO pagos (username, mes, monto, comprobante, created_by)
            VALUES (%s,%s,%s,%s,%s)
            RETURNING id
        """, (username, mes, monto, comprobante, session.get('user')))
        payment_id = c.fetchone()[0]
    else:
        c.execute("""
            INSERT INTO pagos (username, mes, monto, comprobante, created_by)
            VALUES (?,?,?,?,?)
        """, (username, mes, monto, comprobante, session.get('user')))
        payment_id = c.lastrowid
    if vencimiento_nuevo:
        c.execute(qmark("UPDATE clientes SET total_pagado = total_pagado + ?, vencimiento=? WHERE username=?"),
                  (monto, vencimiento_nuevo, username))
    else:
        c.execute(qmark("UPDATE clientes SET total_pagado = total_pagado + ? WHERE username=?"),
                  (monto, username))
    db.commit()
    db.close()
    return jsonify({'ok': True, 'payment_id': payment_id})

@app.route('/api/admin/corregir-mes', methods=['POST'])
@login_required
def corregir_mes():
    """Corrige pagos registrados con mes incorrecto (UTC vs Guatemala)."""
    if session.get('rol') != 'admin':
        return jsonify({'error': 'Sin permiso'}), 403
    data = request.json or {}
    mes_incorrecto = data.get('mes_incorrecto', '2026-05')
    mes_correcto   = data.get('mes_correcto',   '2026-04-30')
    db = get_db()
    c = db.cursor()
    # Buscar pagos con el mes incorrecto
    c.execute(qmark("SELECT id, username, monto, mes, fecha_registro FROM pagos WHERE mes LIKE ?"),
              (f'{mes_incorrecto}%',))
    pagos = fetchall(c)
    if not pagos:
        db.close()
        return jsonify({'ok': True, 'corregidos': 0, 'mensaje': 'No se encontraron pagos con ese mes'})
    ids = [p['id'] for p in pagos]
    # Corregir
    for pid in ids:
        c.execute(qmark("UPDATE pagos SET mes=? WHERE id=?"), (mes_correcto, pid))
    db.commit()
    db.close()
    return jsonify({'ok': True, 'corregidos': len(ids),
                    'pagos': [{'id': p['id'], 'username': p['username'],
                               'monto': p['monto'], 'mes_anterior': p['mes']} for p in pagos]})

@app.route('/api/pagos/<int:pago_id>', methods=['DELETE'])
@login_required
def eliminar_pago(pago_id):
    if session.get('rol') != 'admin':
        return jsonify({'error': 'Acceso denegado'}), 403
    db = get_db()
    c = db.cursor()
    c.execute(qmark("SELECT username, monto FROM pagos WHERE id=?"), (pago_id,))
    pago = fetchone(c)
    if not pago:
        db.close()
        return jsonify({'error': 'Pago no encontrado'}), 404
    c.execute(qmark("DELETE FROM pagos WHERE id=?"), (pago_id,))
    c.execute(qmark("""UPDATE clientes SET total_pagado =
        CASE WHEN total_pagado - ? < 0 THEN 0 ELSE total_pagado - ? END
        WHERE username=?"""), (pago['monto'], pago['monto'], pago['username']))
    db.commit()
    db.close()
    return jsonify({'ok': True, 'username': pago['username'], 'monto': pago['monto']})

@app.route('/api/pagos/<int:pago_id>/comprobante')
@login_required
def get_comprobante(pago_id):
    db = get_db()
    c = db.cursor()
    c.execute(qmark("SELECT comprobante FROM pagos WHERE id=?"), (pago_id,))
    row = fetchone(c)
    db.close()
    if not row or not row['comprobante']:
        return jsonify({'error': 'No encontrado'}), 404
    return jsonify({'comprobante': row['comprobante']})

# ── API: FÉNIX OPERACIÓN / ACTIVACIONES ──────────────────────────────────────
@app.route('/api/plans')
@login_required
def api_plans():
    db = get_db()
    c = db.cursor()
    c.execute("SELECT * FROM plans WHERE is_active = TRUE ORDER BY months ASC, price_gtq ASC NULLS LAST" if PG else
              "SELECT * FROM plans WHERE is_active = 1 ORDER BY months ASC, price_gtq IS NULL, price_gtq ASC")
    rows = fetchall(c)
    db.close()
    return jsonify({'plans': rows})


@app.route('/api/activation-tasks')
@login_required
def api_activation_tasks():
    status = request.args.get('status', 'open')
    db = get_db()
    c = db.cursor()
    where = ""
    params = []
    if status == 'open':
        where = "WHERE t.status IN (?, ?)"
        params = ['pending', 'in_progress']
    elif status:
        where = "WHERE t.status = ?"
        params = [status]
    c.execute(qmark(f"""
        SELECT
          t.*, o.type as order_type, o.amount, o.currency, o.status as order_status,
          o.payment_registered_at, o.payment_id,
          (o.payment_proof IS NOT NULL) as has_payment_proof,
          p.name as plan_name, p.months, p.connections,
          cl.nombre, cl.contacto, cl.vencimiento
        FROM activation_tasks t
        LEFT JOIN orders o ON o.id = t.order_id
        LEFT JOIN plans p ON p.id = o.plan_id
        LEFT JOIN clientes cl ON cl.username = t.username
        {where}
        ORDER BY t.created_at ASC
        LIMIT 100
    """), params)
    rows = fetchall(c)
    c.execute("""
        SELECT status, COUNT(*) as total
        FROM activation_tasks
        GROUP BY status
    """)
    counts_rows = fetchall(c)
    counts = {r['status']: r['total'] for r in counts_rows}
    counts['open'] = counts.get('pending', 0) + counts.get('in_progress', 0)
    db.close()
    return jsonify({'tasks': rows, 'counts': counts, 'status': status})


@app.route('/api/activation-tasks', methods=['POST'])
@login_required
def api_create_activation_task():
    data = request.json or {}
    username = (data.get('username') or '').strip()
    plan_id = data.get('plan_id')
    order_type = data.get('type') or 'renewal'
    amount = float(data.get('amount') or 0)
    currency = data.get('currency') or 'GTQ'
    payment_method = data.get('payment_method') or 'manual'
    payment_proof = data.get('payment_proof') or None
    notes = data.get('notes') or ''
    if not username or not plan_id:
        return jsonify({'error': 'Usuario y plan son obligatorios'}), 400

    db = get_db()
    c = db.cursor()
    c.execute(qmark("SELECT username FROM clientes WHERE username=?"), (username,))
    if not fetchone(c):
        db.close()
        return jsonify({'error': 'Cliente no encontrado'}), 404
    c.execute(qmark("SELECT * FROM plans WHERE id=?"), (plan_id,))
    plan = fetchone(c)
    if not plan:
        db.close()
        return jsonify({'error': 'Plan no encontrado'}), 404

    credits = int(plan['credits_required'] or 0)
    task_type = 'create_line' if order_type == 'new' else 'renew_line'
    if PG:
        c.execute("""
            INSERT INTO orders (username, plan_id, type, status, amount, currency, credits_required, payment_method, payment_proof, notes)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            RETURNING id
        """, (username, plan_id, order_type, 'pending_activation', amount, currency, credits, payment_method, payment_proof, notes))
        order_id = c.fetchone()[0]
        c.execute("""
            INSERT INTO activation_tasks (order_id, username, task_type, status, assigned_to, credits_to_consume, notes)
            VALUES (%s,%s,%s,%s,%s,%s,%s)
            RETURNING id
        """, (order_id, username, task_type, 'pending', session.get('user'), credits, notes))
        task_id = c.fetchone()[0]
    else:
        c.execute("""
            INSERT INTO orders (username, plan_id, type, status, amount, currency, credits_required, payment_method, payment_proof, notes)
            VALUES (?,?,?,?,?,?,?,?,?,?)
        """, (username, plan_id, order_type, 'pending_activation', amount, currency, credits, payment_method, payment_proof, notes))
        order_id = c.lastrowid
        c.execute("""
            INSERT INTO activation_tasks (order_id, username, task_type, status, assigned_to, credits_to_consume, notes)
            VALUES (?,?,?,?,?,?,?)
        """, (order_id, username, task_type, 'pending', session.get('user'), credits, notes))
        task_id = c.lastrowid
    db.commit()
    db.close()
    return jsonify({'ok': True, 'order_id': order_id, 'task_id': task_id})


@app.route('/api/orders/<int:order_id>/payment-proof')
@login_required
def api_order_payment_proof(order_id):
    db = get_db()
    c = db.cursor()
    c.execute(qmark("SELECT payment_proof FROM orders WHERE id=?"), (order_id,))
    row = fetchone(c)
    db.close()
    if not row or not row['payment_proof']:
        return jsonify({'error': 'No encontrado'}), 404
    return jsonify({'payment_proof': row['payment_proof']})


@app.route('/api/activation-tasks/<int:task_id>', methods=['PUT'])
@login_required
def api_update_activation_task(task_id):
    data = request.json or {}
    status = data.get('status')
    allowed = {'pending', 'in_progress', 'done', 'blocked', 'cancelled'}
    if status not in allowed:
        return jsonify({'error': 'Estado inválido'}), 400

    xui_username = data.get('xui_username') or ''
    xui_password = data.get('xui_password') or ''
    xui_expires_at = data.get('xui_expires_at') or None
    portal_password = data.get('portal_password') or ''
    notes = data.get('notes') or ''
    blocked_reason = data.get('blocked_reason') or ''
    register_payment = data.get('register_payment', True)
    maxplayer_sync = bool(data.get('maxplayer_sync', False))
    maxplayer_mode = data.get('maxplayer_mode') or 'create'
    db = get_db()
    c = db.cursor()
    c.execute(qmark("SELECT * FROM activation_tasks WHERE id=?"), (task_id,))
    task = fetchone(c)
    if not task:
        db.close()
        return jsonify({'error': 'Tarea no encontrada'}), 404

    maxplayer_user_id = None
    maxplayer_synced_at = None
    maxplayer_sync_status = None
    if status == 'done' and maxplayer_sync:
        if not xui_username or not xui_password:
            db.close()
            return jsonify({'error': 'Para crear en Max Player se requiere Usuario XUI y Password IPTV.'}), 400
        c.execute(qmark("SELECT nombre, contacto FROM clientes WHERE username=?"), (task['username'],))
        client_row = fetchone(c) or {}
        c.execute(qmark("SELECT maxplayer_user_id FROM client_service_credentials WHERE username=?"), (task['username'],))
        existing_service = fetchone(c) or {}
        try:
            def create_current_maxplayer_user():
                return create_maxplayer_user(
                    username=xui_username,
                    iptv_user=xui_username,
                    iptv_pass=xui_password,
                    password=xui_password,
                    fullname=client_row.get('nombre') or task['username'],
                    user_email=''
                )
            if maxplayer_mode == 'recreate':
                existing_user_id = existing_service.get('maxplayer_user_id')
                if not existing_user_id:
                    existing_user_id = find_maxplayer_user_id(xui_username)
                if not existing_user_id:
                    db.close()
                    return jsonify({'error': f'No encontré el usuario {xui_username} en Max Player para borrarlo. Puedes usar "Crear nuevo" o revisar el username en Max Player.'}), 400
                purge_maxplayer_user(existing_user_id)
            try:
                maxplayer_response, maxplayer_user_id = create_current_maxplayer_user()
            except MaxPlayerError as exc:
                if maxplayer_mode == 'recreate' and is_maxplayer_exists_error(exc):
                    existing_user_id = find_maxplayer_user_id(xui_username)
                    if not existing_user_id:
                        raise
                    purge_maxplayer_user(existing_user_id)
                    maxplayer_response, maxplayer_user_id = create_current_maxplayer_user()
                else:
                    raise
            maxplayer_synced_at = datetime.now(GT_TZ).isoformat()
            maxplayer_sync_status = 'recreated' if maxplayer_mode == 'recreate' else 'created'
            if not maxplayer_user_id:
                maxplayer_sync_status = 'recreated_no_id' if maxplayer_mode == 'recreate' else 'created_no_id'
        except MaxPlayerError as exc:
            db.close()
            return jsonify({'error': str(exc)}), 400

    if status == 'done' and not portal_password:
        portal_password = task['username']

    completed_at = datetime.now(GT_TZ).isoformat() if status == 'done' else None
    if status in {'pending', 'in_progress', 'cancelled'}:
        c.execute(qmark("""
            UPDATE activation_tasks
            SET status=?, assigned_to=?
            WHERE id=?
        """), (status, session.get('user'), task_id))
    else:
        c.execute(qmark("""
            UPDATE activation_tasks
            SET status=?, xui_username=?, xui_password=?, xui_expires_at=?, notes=?, blocked_reason=?, completed_at=?
            WHERE id=?
        """), (status, xui_username, xui_password, xui_expires_at, notes, blocked_reason, completed_at, task_id))
    if task.get('order_id'):
        order_status = 'activated' if status == 'done' else ('blocked' if status == 'blocked' else 'in_activation')
        c.execute(qmark("UPDATE orders SET status=?, completed_at=? WHERE id=?"),
                  (order_status, completed_at, task['order_id']))
    if status == 'done' and xui_expires_at:
        payment_registered = False
        order = None
        if task.get('order_id'):
            c.execute(qmark("SELECT * FROM orders WHERE id=?"), (task['order_id'],))
            order = fetchone(c)
        if register_payment and order and not order.get('payment_registered_at') and float(order.get('amount') or 0) > 0:
            payment_month = today_gt().strftime('%Y-%m-01')
            amount = float(order.get('amount') or 0)
            comprobante = order.get('payment_proof')
            if PG:
                c.execute("""
                    INSERT INTO pagos (username, mes, monto, comprobante)
                    VALUES (%s,%s,%s,%s)
                    RETURNING id
                """, (task['username'], payment_month, amount, comprobante))
                payment_id = c.fetchone()[0]
            else:
                c.execute("""
                    INSERT INTO pagos (username, mes, monto, comprobante)
                    VALUES (?,?,?,?)
                """, (task['username'], payment_month, amount, comprobante))
                payment_id = c.lastrowid
            c.execute(qmark("UPDATE orders SET payment_registered_at=?, payment_id=? WHERE id=?"),
                      (completed_at, payment_id, task['order_id']))
            c.execute(qmark("UPDATE clientes SET total_pagado = total_pagado + ?, vencimiento=? WHERE username=?"),
                      (amount, xui_expires_at, task['username']))
            payment_registered = True
        if not payment_registered:
            c.execute(qmark("UPDATE clientes SET vencimiento=? WHERE username=?"),
                      (xui_expires_at, task['username']))
        if PG:
            c.execute("""
                INSERT INTO client_service_credentials
                    (username, app_name, service_username, service_password, expires_at, devices,
                     maxplayer_user_id, maxplayer_synced_at, maxplayer_sync_status, updated_at)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                ON CONFLICT (username) DO UPDATE SET
                    app_name=EXCLUDED.app_name,
                    service_username=EXCLUDED.service_username,
                    service_password=CASE
                        WHEN EXCLUDED.service_password IS NULL OR EXCLUDED.service_password = ''
                        THEN client_service_credentials.service_password
                        ELSE EXCLUDED.service_password
                    END,
                    expires_at=EXCLUDED.expires_at,
                    devices=EXCLUDED.devices,
                    maxplayer_user_id=COALESCE(EXCLUDED.maxplayer_user_id, client_service_credentials.maxplayer_user_id),
                    maxplayer_synced_at=COALESCE(EXCLUDED.maxplayer_synced_at, client_service_credentials.maxplayer_synced_at),
                    maxplayer_sync_status=COALESCE(EXCLUDED.maxplayer_sync_status, client_service_credentials.maxplayer_sync_status),
                    updated_at=EXCLUDED.updated_at
            """, (task['username'], 'Max Player', xui_username, xui_password, xui_expires_at, 3,
                  maxplayer_user_id, maxplayer_synced_at, maxplayer_sync_status, completed_at))
            if portal_password:
                c.execute("""
                    INSERT INTO client_portal_accounts (username, password, is_enabled, updated_at)
                    VALUES (%s,%s,TRUE,%s)
                    ON CONFLICT (username) DO UPDATE SET
                        password=EXCLUDED.password,
                        is_enabled=TRUE,
                        updated_at=EXCLUDED.updated_at
                """, (task['username'], hash_password(portal_password), completed_at))
        else:
            c.execute("""
                INSERT INTO client_service_credentials
                    (username, app_name, service_username, service_password, expires_at, devices,
                     maxplayer_user_id, maxplayer_synced_at, maxplayer_sync_status, updated_at)
                VALUES (?,?,?,?,?,?,?,?,?,?)
                ON CONFLICT(username) DO UPDATE SET
                    app_name=excluded.app_name,
                    service_username=excluded.service_username,
                    service_password=CASE
                        WHEN excluded.service_password IS NULL OR excluded.service_password = ''
                        THEN client_service_credentials.service_password
                        ELSE excluded.service_password
                    END,
                    expires_at=excluded.expires_at,
                    devices=excluded.devices,
                    maxplayer_user_id=COALESCE(excluded.maxplayer_user_id, client_service_credentials.maxplayer_user_id),
                    maxplayer_synced_at=COALESCE(excluded.maxplayer_synced_at, client_service_credentials.maxplayer_synced_at),
                    maxplayer_sync_status=COALESCE(excluded.maxplayer_sync_status, client_service_credentials.maxplayer_sync_status),
                    updated_at=excluded.updated_at
            """, (task['username'], 'Max Player', xui_username, xui_password, xui_expires_at, 3,
                  maxplayer_user_id, maxplayer_synced_at, maxplayer_sync_status, completed_at))
            if portal_password:
                c.execute("""
                    INSERT INTO client_portal_accounts (username, password, is_enabled, updated_at)
                    VALUES (?,?,1,?)
                    ON CONFLICT(username) DO UPDATE SET
                        password=excluded.password,
                        is_enabled=1,
                        updated_at=excluded.updated_at
                """, (task['username'], hash_password(portal_password), completed_at))
    db.commit()
    db.close()
    return jsonify({'ok': True})

# ── API: ANALYTICS ────────────────────────────────────────────────────────────
@app.route('/api/activation-tasks/<int:task_id>', methods=['DELETE'])
@login_required
def api_delete_activation_task(task_id):
    if session.get('rol') != 'admin':
        return jsonify({'error': 'Acceso denegado'}), 403
    db = get_db()
    c = db.cursor()
    c.execute(qmark("SELECT * FROM activation_tasks WHERE id=?"), (task_id,))
    task = fetchone(c)
    if not task:
        db.close()
        return jsonify({'error': 'Tarea no encontrada'}), 404

    pago_revertido = False
    if task.get('order_id'):
        c.execute(qmark("SELECT * FROM orders WHERE id=?"), (task['order_id'],))
        order = fetchone(c)
        if order and order.get('payment_id'):
            c.execute(qmark("SELECT * FROM pagos WHERE id=?"), (order['payment_id'],))
            pago = fetchone(c)
            if pago:
                c.execute(qmark("DELETE FROM pagos WHERE id=?"), (order['payment_id'],))
                if PG:
                    c.execute("UPDATE clientes SET total_pagado = GREATEST(total_pagado - %s, 0) WHERE username=%s",
                              (float(pago.get('monto') or 0), pago['username']))
                else:
                    c.execute("UPDATE clientes SET total_pagado = MAX(total_pagado - ?, 0) WHERE username=?",
                              (float(pago.get('monto') or 0), pago['username']))
                pago_revertido = True

    c.execute(qmark("DELETE FROM activation_tasks WHERE id=?"), (task_id,))
    if task.get('order_id'):
        c.execute(qmark("DELETE FROM orders WHERE id=?"), (task['order_id'],))
    db.commit()
    db.close()
    return jsonify({'ok': True, 'pago_revertido': pago_revertido})


@app.route('/api/analytics')
@login_required
def analytics():
    if session.get('rol') != 'admin':
        return jsonify({'error': 'Acceso denegado'}), 403

    db = get_db()
    c = db.cursor()
    credit_cost_usd = CREDIT_COST_USD
    usd_gtq_rate = USD_GTQ_RATE
    credit_cost_gtq = credit_cost_usd * usd_gtq_rate

    # Ventas mensuales (todo el tiempo)
    if PG:
        c.execute("""
            SELECT TO_CHAR(mes::date, 'YYYY-MM') as m,
                   SUM(monto) as total, COUNT(*) as n_pagos,
                   COUNT(DISTINCT username) as clientes
            FROM pagos GROUP BY m ORDER BY m
        """)
    else:
        c.execute("""
            SELECT strftime('%Y-%m', mes) as m,
                   SUM(monto) as total, COUNT(*) as n_pagos,
                   COUNT(DISTINCT username) as clientes
            FROM pagos GROUP BY m ORDER BY m
        """)
    ventas_mensuales = fetchall(c)

    # Ventas anuales
    if PG:
        c.execute("""
            SELECT TO_CHAR(mes::date, 'YYYY') as y,
                   SUM(monto) as total, COUNT(DISTINCT username) as clientes
            FROM pagos GROUP BY y ORDER BY y
        """)
    else:
        c.execute("""
            SELECT strftime('%Y', mes) as y,
                   SUM(monto) as total, COUNT(DISTINCT username) as clientes
            FROM pagos GROUP BY y ORDER BY y
        """)
    ventas_anuales = fetchall(c)

    # Distribución por tipo de plan (según monto del pago)
    if PG:
        c.execute("""
            SELECT
                CASE
                    WHEN monto <= 120  THEN 'Mensual'
                    WHEN monto <= 320  THEN 'Trimestral'
                    WHEN monto <= 650  THEN 'Semestral'
                    WHEN monto <= 1300 THEN 'Anual'
                    ELSE 'Más de 1 año'
                END as plan,
                COUNT(*) as n, SUM(monto) as total
            FROM pagos GROUP BY plan ORDER BY n DESC
        """)
    else:
        c.execute("""
            SELECT
                CASE
                    WHEN monto <= 120  THEN 'Mensual'
                    WHEN monto <= 320  THEN 'Trimestral'
                    WHEN monto <= 650  THEN 'Semestral'
                    WHEN monto <= 1300 THEN 'Anual'
                    ELSE 'Más de 1 año'
                END as plan,
                COUNT(*) as n, SUM(monto) as total
            FROM pagos GROUP BY plan ORDER BY n DESC
        """)
    planes = fetchall(c)

    # Clientes nuevos por mes (primer pago de cada usuario)
    if PG:
        c.execute("""
            SELECT TO_CHAR(p.mes::date, 'YYYY-MM') as m, COUNT(*) as nuevos
            FROM pagos p
            WHERE p.mes = (SELECT MIN(p2.mes) FROM pagos p2 WHERE p2.username = p.username)
            GROUP BY m ORDER BY m
        """)
    else:
        c.execute("""
            SELECT strftime('%Y-%m', p.mes) as m, COUNT(*) as nuevos
            FROM pagos p
            WHERE p.mes = (SELECT MIN(p2.mes) FROM pagos p2 WHERE p2.username = p.username)
            GROUP BY m ORDER BY m
        """)
    nuevos_por_mes = fetchall(c)

    # Renovaciones por mes (pagos que NO son el primero del usuario)
    if PG:
        c.execute("""
            SELECT TO_CHAR(p.mes::date, 'YYYY-MM') as m, COUNT(*) as renovaciones
            FROM pagos p
            WHERE p.mes != (SELECT MIN(p2.mes) FROM pagos p2 WHERE p2.username = p.username)
            GROUP BY m ORDER BY m
        """)
    else:
        c.execute("""
            SELECT strftime('%Y-%m', p.mes) as m, COUNT(*) as renovaciones
            FROM pagos p
            WHERE p.mes != (SELECT MIN(p2.mes) FROM pagos p2 WHERE p2.username = p.username)
            GROUP BY m ORDER BY m
        """)
    renovaciones_por_mes = fetchall(c)

    # Clientes que no renovaron (vencidos con pagos históricos)
    today = today_gt().isoformat()
    c.execute(qmark("SELECT COUNT(*) as c FROM clientes WHERE (vencimiento < ? OR vencimiento IS NULL) AND total_pagado > 0"), (today,))
    no_renovaron = fetchone(c)['c']

    # Total histórico
    c.execute("SELECT COALESCE(SUM(monto),0) as t FROM pagos")
    total_historico = round(float(fetchone(c)['t']), 2)

    c.execute("SELECT COUNT(*) as c FROM clientes")
    total_clientes = fetchone(c)['c']

    # Ventas por día (últimos 60 días)
    mes_actual = today_gt().strftime('%Y-%m')
    mes_anterior = (today_gt().replace(day=1) - timedelta(days=1)).strftime('%Y-%m')
    if PG:
        c.execute("""
            SELECT TO_CHAR(mes::date, 'YYYY-MM-DD') as dia,
                   SUM(monto) as total, COUNT(*) as n_pagos
            FROM pagos
            WHERE SUBSTRING(mes,1,7) IN (%s, %s)
            GROUP BY dia ORDER BY dia
        """, (mes_actual, mes_anterior))
    else:
        c.execute("""
            SELECT strftime('%Y-%m-%d', mes) as dia,
                   SUM(monto) as total, COUNT(*) as n_pagos
            FROM pagos
            WHERE substr(mes,1,7) IN (?, ?)
            GROUP BY dia ORDER BY dia
        """, (mes_actual, mes_anterior))
    ventas_por_dia = fetchall(c)

    # Créditos históricos estimados desde pagos antiguos; no modifica historial.
    c.execute("SELECT COALESCE(SUM(monto),0) as ingresos, COUNT(*) as pagos FROM pagos")
    retro_base = fetchone(c)
    c.execute("SELECT monto FROM pagos")
    retro_pagos = fetchall(c)
    retro_creditos = sum(estimate_credits_from_amount(p.get('monto')) for p in retro_pagos)
    retro_ingresos = float(retro_base['ingresos'] or 0)
    retro_costo = retro_creditos * credit_cost_gtq
    retro_utilidad = retro_ingresos - retro_costo
    retro_margen = round((retro_utilidad / retro_ingresos) * 100, 1) if retro_ingresos > 0 else 0

    # Operación Fénix: créditos, utilidad y tareas (mes actual)
    if PG:
        c.execute("""
            SELECT
                COALESCE(SUM(o.amount),0) as ingresos,
                COALESCE(SUM(t.credits_to_consume),0) as creditos,
                COUNT(*) as completadas
            FROM activation_tasks t
            LEFT JOIN orders o ON o.id = t.order_id
            WHERE t.status = 'done'
              AND SUBSTRING(COALESCE(t.completed_at, ''), 1, 7) = %s
        """, (mes_actual,))
    else:
        c.execute("""
            SELECT
                COALESCE(SUM(o.amount),0) as ingresos,
                COALESCE(SUM(t.credits_to_consume),0) as creditos,
                COUNT(*) as completadas
            FROM activation_tasks t
            LEFT JOIN orders o ON o.id = t.order_id
            WHERE t.status = 'done'
              AND substr(COALESCE(t.completed_at, ''), 1, 7) = ?
        """, (mes_actual,))
    op_mes = fetchone(c)
    op_ingresos = float(op_mes['ingresos'] or 0)
    op_creditos = int(op_mes['creditos'] or 0)
    op_costo = op_creditos * credit_cost_gtq
    op_utilidad = op_ingresos - op_costo
    op_margin = round((op_utilidad / op_ingresos) * 100, 1) if op_ingresos > 0 else 0

    c.execute("""
        SELECT status, COUNT(*) as total
        FROM activation_tasks
        GROUP BY status
        ORDER BY status
    """)
    tareas_por_estado = fetchall(c)

    if PG:
        c.execute("""
            SELECT
                COALESCE(assigned_to, 'Sin asignar') as agente,
                COUNT(*) as tareas,
                COALESCE(SUM(CASE WHEN t.status = 'done' THEN t.credits_to_consume ELSE 0 END),0) as creditos,
                COALESCE(SUM(CASE WHEN t.status = 'done' THEN o.amount ELSE 0 END),0) as ingresos
            FROM activation_tasks t
            LEFT JOIN orders o ON o.id = t.order_id
            WHERE SUBSTRING(COALESCE(t.created_at, ''), 1, 7) = %s
            GROUP BY COALESCE(assigned_to, 'Sin asignar')
            ORDER BY tareas DESC
            LIMIT 10
        """, (mes_actual,))
    else:
        c.execute("""
            SELECT
                COALESCE(assigned_to, 'Sin asignar') as agente,
                COUNT(*) as tareas,
                COALESCE(SUM(CASE WHEN t.status = 'done' THEN t.credits_to_consume ELSE 0 END),0) as creditos,
                COALESCE(SUM(CASE WHEN t.status = 'done' THEN o.amount ELSE 0 END),0) as ingresos
            FROM activation_tasks t
            LEFT JOIN orders o ON o.id = t.order_id
            WHERE substr(COALESCE(t.created_at, ''), 1, 7) = ?
            GROUP BY COALESCE(assigned_to, 'Sin asignar')
            ORDER BY tareas DESC
            LIMIT 10
        """, (mes_actual,))
    tareas_por_agente = fetchall(c)

    if PG:
        c.execute("""
            SELECT
                TO_CHAR(t.completed_at::timestamp, 'YYYY-MM') as m,
                COALESCE(SUM(t.credits_to_consume),0) as creditos,
                COALESCE(SUM(o.amount),0) as ingresos,
                COUNT(*) as completadas
            FROM activation_tasks t
            LEFT JOIN orders o ON o.id = t.order_id
            WHERE t.status = 'done' AND t.completed_at IS NOT NULL
            GROUP BY m
            ORDER BY m
        """)
    else:
        c.execute("""
            SELECT
                strftime('%Y-%m', t.completed_at) as m,
                COALESCE(SUM(t.credits_to_consume),0) as creditos,
                COALESCE(SUM(o.amount),0) as ingresos,
                COUNT(*) as completadas
            FROM activation_tasks t
            LEFT JOIN orders o ON o.id = t.order_id
            WHERE t.status = 'done' AND t.completed_at IS NOT NULL
            GROUP BY m
            ORDER BY m
        """)
    operacion_mensual = fetchall(c)

    db.close()

    return jsonify({
        'ventas_mensuales': ventas_mensuales,
        'ventas_anuales': ventas_anuales,
        'planes': planes,
        'nuevos_por_mes': nuevos_por_mes,
        'renovaciones_por_mes': renovaciones_por_mes,
        'no_renovaron': no_renovaron,
        'total_historico': total_historico,
        'total_clientes': total_clientes,
        'ventas_por_dia': ventas_por_dia,
        'operacion': {
            'mes': mes_actual,
            'credit_cost_usd': credit_cost_usd,
            'usd_gtq_rate': usd_gtq_rate,
            'credit_cost_gtq': credit_cost_gtq,
            'ingresos_mes': round(op_ingresos, 2),
            'creditos_mes': op_creditos,
            'costo_mes': round(op_costo, 2),
            'utilidad_mes': round(op_utilidad, 2),
            'margen_mes': op_margin,
            'completadas_mes': int(op_mes['completadas'] or 0),
            'tareas_por_estado': tareas_por_estado,
            'tareas_por_agente': tareas_por_agente,
            'operacion_mensual': operacion_mensual,
            'retro': {
                'ingresos': round(retro_ingresos, 2),
                'pagos': int(retro_base['pagos'] or 0),
                'creditos_estimados': int(retro_creditos),
                'costo_estimado': round(retro_costo, 2),
                'utilidad_estimada': round(retro_utilidad, 2),
                'margen_estimado': retro_margen
            }
        }
    })


# ── CLIENTES NUEVOS POR RANGO ─────────────────────────────────────────────────
@app.route('/api/admin/clientes-nuevos')
@login_required
def clientes_nuevos():
    if session.get('rol') != 'admin':
        return jsonify({'error': 'Acceso denegado'}), 403
    desde = request.args.get('desde', '')
    hasta = request.args.get('hasta', '')
    if not desde or not hasta:
        return jsonify({'error': 'Parámetros desde y hasta requeridos'}), 400
    db = get_db()
    c = db.cursor()
    if PG:
        c.execute("""
            SELECT cl.username, cl.nombre, cl.contacto, cl.vencimiento,
                   cl.total_pagado, cl.created_at,
                   COUNT(p.id) as num_pagos,
                   MAX(p.mes) as ultimo_pago
            FROM clientes cl
            LEFT JOIN pagos p ON p.username = cl.username
            WHERE SUBSTRING(cl.created_at, 1, 10) >= %s
              AND SUBSTRING(cl.created_at, 1, 10) <= %s
            GROUP BY cl.username, cl.nombre, cl.contacto,
                     cl.vencimiento, cl.total_pagado, cl.created_at
            ORDER BY cl.created_at ASC
        """, (desde, hasta))
    else:
        c.execute("""
            SELECT cl.username, cl.nombre, cl.contacto, cl.vencimiento,
                   cl.total_pagado, cl.created_at,
                   COUNT(p.id) as num_pagos,
                   MAX(p.mes) as ultimo_pago
            FROM clientes cl
            LEFT JOIN pagos p ON p.username = cl.username
            WHERE substr(cl.created_at, 1, 10) >= ?
              AND substr(cl.created_at, 1, 10) <= ?
            GROUP BY cl.username, cl.nombre, cl.contacto,
                     cl.vencimiento, cl.total_pagado, cl.created_at
            ORDER BY cl.created_at ASC
        """, (desde, hasta))
    rows = fetchall(c)
    db.close()
    return jsonify({'clientes': rows, 'total': len(rows), 'desde': desde, 'hasta': hasta})


# ── MIGRACIÓN MASIVA ─────────────────────────────────────────────────────────
@app.route('/api/admin/migracion-clientes')
@login_required
def migracion_clientes():
    q = request.args.get('q', '').strip()
    estado = request.args.get('estado', '').strip()
    today = today_gt().isoformat()
    params = []
    where = []
    if q:
        like = f'%{q}%'
        where.append("(cl.username LIKE ? OR cl.nombre LIKE ? OR cl.contacto LIKE ? OR svc.service_username LIKE ?)")
        params.extend([like, like, like, like])
    if estado == 'activos':
        where.append("cl.vencimiento >= ?")
        params.append(today)
    elif estado == 'vencidos':
        where.append("(cl.vencimiento < ? OR cl.vencimiento IS NULL OR cl.vencimiento = '')")
        params.append(today)
    where_sql = ('WHERE ' + ' AND '.join(where)) if where else ''

    db = get_db()
    c = db.cursor()
    c.execute(qmark(f"""
        SELECT
            cl.username, cl.nombre, cl.contacto, cl.vencimiento,
            acc.username AS portal_username,
            acc.is_enabled AS portal_enabled,
            svc.service_username, svc.service_password, svc.maxplayer_user_id,
            svc.maxplayer_synced_at, svc.maxplayer_sync_status
        FROM clientes cl
        LEFT JOIN client_portal_accounts acc ON acc.username = cl.username
        LEFT JOIN client_service_credentials svc ON svc.username = cl.username
        {where_sql}
        ORDER BY
            CASE WHEN cl.vencimiento IS NULL OR cl.vencimiento = '' THEN 1 ELSE 0 END,
            cl.vencimiento DESC,
            cl.username ASC
        LIMIT 500
    """), params)
    rows = fetchall(c)
    db.close()

    items = []
    counts = {'total': 0, 'sin_portal': 0, 'sin_xui': 0, 'maxplayer_pendiente': 0, 'maxplayer_listo': 0}
    for r in rows:
        portal_ok = bool(r.get('portal_username')) and bool(r.get('portal_enabled'))
        service_username = r.get('service_username') or ''
        service_password = r.get('service_password') or ''
        maxplayer_ok = bool(r.get('maxplayer_user_id')) and (
            r.get('maxplayer_sync_status') in {'restored', 'created', 'done'} or bool(r.get('maxplayer_synced_at'))
        )
        counts['total'] += 1
        if not portal_ok:
            counts['sin_portal'] += 1
        if not service_username or not service_password:
            counts['sin_xui'] += 1
        if maxplayer_ok:
            counts['maxplayer_listo'] += 1
        else:
            counts['maxplayer_pendiente'] += 1
        items.append({
            'username': r.get('username'),
            'nombre': r.get('nombre'),
            'contacto': r.get('contacto'),
            'vencimiento': r.get('vencimiento'),
            'portal_ok': portal_ok,
            'service_username': service_username,
            'has_service_password': bool(service_password),
            'maxplayer_ok': maxplayer_ok,
            'maxplayer_user_id': r.get('maxplayer_user_id'),
            'maxplayer_sync_status': r.get('maxplayer_sync_status') or '',
            'maxplayer_synced_at': r.get('maxplayer_synced_at') or '',
        })
    return jsonify({'clientes': items, 'counts': counts})


# ── EXPORTAR CLIENTES (datos JSON para generar Excel en el frontend) ──────────
@app.route('/api/admin/exportar-clientes')
@login_required
def exportar_clientes():
    if session.get('rol') != 'admin':
        return jsonify({'error': 'Acceso denegado'}), 403
    db = get_db()
    c = db.cursor()
    if PG:
        c.execute("""
            SELECT cl.username, cl.nombre, cl.contacto, cl.vencimiento,
                   cl.total_pagado,
                   MAX(p.mes) as ultimo_pago,
                   COUNT(p.id) as num_pagos
            FROM clientes cl
            LEFT JOIN pagos p ON p.username = cl.username
            GROUP BY cl.username, cl.nombre, cl.contacto, cl.vencimiento, cl.total_pagado
            ORDER BY cl.vencimiento DESC NULLS LAST, cl.username ASC
        """)
    else:
        c.execute("""
            SELECT cl.username, cl.nombre, cl.contacto, cl.vencimiento,
                   cl.total_pagado,
                   MAX(p.mes) as ultimo_pago,
                   COUNT(p.id) as num_pagos
            FROM clientes cl
            LEFT JOIN pagos p ON p.username = cl.username
            GROUP BY cl.username, cl.nombre, cl.contacto, cl.vencimiento, cl.total_pagado
            ORDER BY CASE WHEN cl.vencimiento IS NULL THEN 1 ELSE 0 END,
                     cl.vencimiento DESC, cl.username ASC
        """)
    rows = fetchall(c)
    db.close()
    return jsonify({'clientes': rows})

# ── CONFIGURACIÓN ─────────────────────────────────────────────────────────────
@app.route('/api/config', methods=['GET'])
@login_required
def get_config():
    db = get_db()
    c = db.cursor()
    c.execute("SELECT clave, valor FROM configuracion")
    rows = fetchall(c)
    db.close()
    return jsonify({r['clave']: r['valor'] for r in rows})


@app.route('/api/config', methods=['POST'])
@login_required
def save_config():
    if session.get('rol') != 'admin':
        return jsonify({'error': 'Acceso denegado'}), 403
    data = request.get_json()
    db = get_db()
    c = db.cursor()
    for clave, valor in data.items():
        if PG:
            c.execute(
                "INSERT INTO configuracion (clave, valor) VALUES (%s, %s) ON CONFLICT (clave) DO UPDATE SET valor = EXCLUDED.valor",
                (clave, valor)
            )
        else:
            c.execute("INSERT OR REPLACE INTO configuracion (clave, valor) VALUES (?, ?)", (clave, valor))
    db.commit()
    db.close()
    return jsonify({'ok': True})


@app.route('/api/cambiar-password', methods=['POST'])
@login_required
def cambiar_password():
    data = request.json
    password_actual = data.get('password_actual', '')
    password_nueva = data.get('password_nueva', '')
    if not password_actual or not password_nueva or len(password_nueva) < 4:
        return jsonify({'error': 'Datos inválidos. La nueva contraseña debe tener al menos 4 caracteres.'}), 400
    username = session['user']
    db = get_db()
    c = db.cursor()
    c.execute(qmark("SELECT id, password FROM usuarios WHERE username=?"), (username,))
    user = fetchone(c)
    if not user or not verify_password(user.get('password'), password_actual):
        db.close()
        return jsonify({'error': 'La contraseña actual es incorrecta.'}), 403
    c.execute(qmark("UPDATE usuarios SET password=? WHERE username=?"), (hash_password(password_nueva), username))
    db.commit()
    db.close()
    return jsonify({'ok': True})


@app.route('/api/admin/reimportar-excel', methods=['POST'])
@login_required
def reimportar_excel():
    if session.get('rol') != 'admin':
        return jsonify({'error': 'Acceso denegado'}), 403
    import tempfile
    usar_tmp = False
    if 'archivo' in request.files:
        f = request.files['archivo']
        suffix = '.xlsx' if f.filename.lower().endswith('.xlsx') else '.xls'
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
        f.save(tmp.name)
        excel_path = tmp.name
        usar_tmp = True
    else:
        excel_path = os.path.join(os.path.dirname(__file__), 'IPTV Nuevo (2).xlsx')
    if not os.path.exists(excel_path):
        return jsonify({'ok': False, 'error': 'Excel no encontrado. Sube el archivo directamente.'})
    try:
        clientes, pagos = _import_excel_rows(excel_path, update_existing=True)
        return jsonify({'ok': True, 'clientes_actualizados': clientes, 'pagos_procesados': pagos})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})
    finally:
        if usar_tmp and os.path.exists(excel_path):
            os.unlink(excel_path)

@app.route('/api/admin/importar-xui-credenciales', methods=['POST'])
@login_required
def importar_xui_credenciales():
    if session.get('rol') != 'admin':
        return jsonify({'error': 'Acceso denegado'}), 403
    if 'archivo' not in request.files:
        return jsonify({'ok': False, 'error': 'Sube un archivo Excel con columnas Username, Password y Vencimiento.'}), 400
    import tempfile
    try:
        from openpyxl import load_workbook
    except ImportError:
        import subprocess, sys
        subprocess.check_call([sys.executable, '-m', 'pip', 'install', '--quiet', 'openpyxl'])
        from openpyxl import load_workbook

    f = request.files['archivo']
    suffix = '.xlsx' if f.filename.lower().endswith('.xlsx') else '.xls'
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
    f.save(tmp.name)
    excel_path = tmp.name
    processed = matched = created_clients = portal_created = 0
    try:
        wb = load_workbook(excel_path, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        if not rows:
            return jsonify({'ok': False, 'error': 'El Excel está vacío.'}), 400
        headers = [str(h or '').strip().lower() for h in rows[0]]
        def col(*names):
            for name in names:
                key = name.lower()
                if key in headers:
                    return headers.index(key)
            return None
        user_col = col('username', 'usuario', 'user')
        pass_col = col('password', 'contraseña', 'contrasena', 'pass')
        exp_col = col('vencimiento', 'expires', 'expiration', 'fecha vencimiento')
        if user_col is None or pass_col is None:
            return jsonify({'ok': False, 'error': 'No encontré columnas Username y Password.'}), 400

        db = get_db()
        c = db.cursor()
        now = datetime.now(GT_TZ).isoformat()
        for row in rows[1:]:
            username = str(row[user_col] or '').strip() if len(row) > user_col else ''
            password = str(row[pass_col] or '').strip() if len(row) > pass_col else ''
            if not username or not password:
                continue
            processed += 1
            expires_at = None
            if exp_col is not None and len(row) > exp_col and row[exp_col]:
                value = row[exp_col]
                if hasattr(value, 'date'):
                    expires_at = value.date().isoformat()
                else:
                    expires_at = str(value).strip()
            c.execute(qmark("SELECT username FROM clientes WHERE username=?"), (username,))
            client_exists = bool(fetchone(c))
            if not client_exists:
                if PG:
                    c.execute("""
                        INSERT INTO clientes
                            (username, nombre, contacto, vencimiento, referido, notas, total_pagado)
                        VALUES (%s,%s,%s,%s,%s,%s,0)
                        ON CONFLICT (username) DO NOTHING
                    """, (username, username, '', expires_at, 'NO', 'Creado automáticamente desde importación XUI'))
                else:
                    c.execute("""
                        INSERT OR IGNORE INTO clientes
                            (username, nombre, contacto, vencimiento, referido, notas, total_pagado)
                        VALUES (?,?,?,?,?,?,0)
                    """, (username, username, '', expires_at, 'NO', 'Creado automáticamente desde importación XUI'))
                created_clients += 1
            matched += 1
            if expires_at:
                c.execute(qmark("UPDATE clientes SET vencimiento=? WHERE username=?"), (expires_at, username))
            if PG:
                c.execute("""
                    INSERT INTO client_service_credentials
                        (username, app_name, service_username, service_password, expires_at, devices, updated_at)
                    VALUES (%s,%s,%s,%s,%s,%s,%s)
                    ON CONFLICT (username) DO UPDATE SET
                        app_name=EXCLUDED.app_name,
                        service_username=EXCLUDED.service_username,
                        service_password=EXCLUDED.service_password,
                        expires_at=COALESCE(EXCLUDED.expires_at, client_service_credentials.expires_at),
                        devices=EXCLUDED.devices,
                        updated_at=EXCLUDED.updated_at
                """, (username, 'Max Player', username, password, expires_at, 3, now))
                c.execute("""
                    INSERT INTO client_portal_accounts (username, password, is_enabled, updated_at)
                    VALUES (%s,%s,TRUE,%s)
                    ON CONFLICT (username) DO NOTHING
                """, (username, hash_password(username), now))
            else:
                c.execute("""
                    INSERT INTO client_service_credentials
                        (username, app_name, service_username, service_password, expires_at, devices, updated_at)
                    VALUES (?,?,?,?,?,?,?)
                    ON CONFLICT(username) DO UPDATE SET
                        app_name=excluded.app_name,
                        service_username=excluded.service_username,
                        service_password=excluded.service_password,
                        expires_at=COALESCE(excluded.expires_at, client_service_credentials.expires_at),
                        devices=excluded.devices,
                        updated_at=excluded.updated_at
                """, (username, 'Max Player', username, password, expires_at, 3, now))
                c.execute("""
                    INSERT OR IGNORE INTO client_portal_accounts (username, password, is_enabled, updated_at)
                    VALUES (?,?,1,?)
                """, (username, hash_password(username), now))
            if c.rowcount:
                portal_created += 1
        db.commit()
        db.close()
        return jsonify({
            'ok': True,
            'filas_procesadas': processed,
            'clientes_actualizados': matched,
            'clientes_creados': created_clients,
            'portal_creados_si_faltaban': portal_created,
            'no_encontrados': [],
            'no_encontrados_total': 0
        })
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500
    finally:
        if os.path.exists(excel_path):
            os.unlink(excel_path)


@app.route('/api/diagnostico')
@login_required
def diagnostico():
    if session.get('rol') != 'admin':
        return jsonify({'error': 'Acceso denegado'}), 403
    db = get_db()
    c = db.cursor()
    # Total pagos
    c.execute("SELECT COUNT(*) as total FROM pagos")
    total = fetchone(c)['total']
    # Últimos 10 pagos registrados
    c.execute(qmark("SELECT id, username, mes, monto, fecha_registro FROM pagos ORDER BY id DESC LIMIT 10"))
    ultimos = fetchall(c)
    # Tablas existentes
    if PG:
        c.execute("SELECT table_name FROM information_schema.tables WHERE table_schema='public'")
    else:
        c.execute("SELECT name FROM sqlite_master WHERE type='table'")
    tablas = [r[0] for r in c.fetchall()]
    db.close()
    return jsonify({'total_pagos': total, 'ultimos_pagos': ultimos, 'tablas': tablas})


if __name__ == '__main__':
    init_db()
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)
